#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Registro de Negocios — Aplicación de escritorio (v3, CRM local)
------------------------------------------------------------------
Guarda toda la base de datos en un archivo local (negocios.db, SQLite)
ubicado en la misma carpeta que este programa. Los datos NO se pierden
al cerrar el programa.

Novedades de esta versión (v3):
    - Buscador (nombre, código, empresa, ciudad, etiquetas).
    - Filtros por ciudad y estado, orden por nombre/código/fecha.
    - Favoritos y etiquetas personalizadas.
    - Selección múltiple: edición masiva y eliminación en lote.
    - Duplicar un negocio (evita choques de código automáticamente).
    - Historial de modificaciones (auditoría por registro y global).
    - Seguimiento comercial: último contacto / próximo seguimiento
      (con selector de calendario propio, sin dependencias extra).
    - Campos nuevos: página web, cargo de contacto, RUC/NIT/CIF.
    - Acciones rápidas: copiar teléfono/correo, abrir WhatsApp,
      redactar en Gmail, abrir sitio web (clic derecho sobre la fila).
    - Importar desde Excel/CSV (fusiona con lo que ya tengas).
    - Exportar a Excel, CSV y PDF.
    - Panel de Dashboard con resumen general.

Requisitos:
    - Python 3.8+ (tkinter y sqlite3 vienen incluidos)
    - openpyxl   -> pip install openpyxl    (Excel)
    - reportlab  -> pip install reportlab   (solo para exportar PDF; opcional)

Ejecutar:
    python app.py
"""

import os
import sys
import csv
import sqlite3
import datetime
import calendar as calendar_module
import unicodedata
import webbrowser
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# ---------------------------------------------------------------------------
# Rutas y base de datos
# ---------------------------------------------------------------------------

APP_DIR = os.path.dirname(os.path.abspath(sys.executable)) if getattr(sys, "frozen", False) \
    else os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "negocios.db")

ESTADOS = ["Activo", "Potencial", "Inactivo"]

# Columnas visibles en las tablas (rubro y vista general)
COLUMNS = [
    ("fav", "★", 34),
    ("code", "Código", 85),
    ("negocio", "Negocio", 170),
    ("contacto", "Contacto", 120),
    ("telefono", "Teléfono", 110),
    ("ciudad", "Ciudad", 90),
    ("estado", "Estado", 85),
    ("ultimo_contacto", "Últ. contacto", 95),
    ("proximo_contacto", "Próx. seguim.", 95),
    ("tags", "Etiquetas", 110),
]
GENERAL_COLUMNS = [("category", "Rubro", 110)] + COLUMNS

# Campos completos de un negocio (usados en formularios / import / export)
FULL_FIELDS = ["negocio", "contacto", "telefono", "correo", "direccion", "ciudad", "estado",
               "notas", "pagina_web", "cargo_contacto", "ruc", "tags",
               "ultimo_contacto", "proximo_contacto"]

FIELD_LABELS = {
    "negocio": "Nombre del negocio", "contacto": "Nombre de contacto",
    "telefono": "Número de contacto", "correo": "Correo electrónico",
    "direccion": "Dirección", "ciudad": "Ciudad", "estado": "Estado",
    "notas": "Notas", "pagina_web": "Página web", "cargo_contacto": "Cargo de contacto",
    "ruc": "RUC/NIT/CIF", "tags": "Etiquetas", "ultimo_contacto": "Último contacto",
    "proximo_contacto": "Próximo seguimiento", "favorito": "Favorito",
}

# Alias aceptados al importar archivos externos (Excel/CSV)
HEADER_ALIASES = {
    "rubro": "rubro", "categoria": "rubro", "category": "rubro", "giro": "rubro",
    "codigo": "code", "code": "code",
    "nombre del negocio": "negocio", "negocio": "negocio", "empresa": "negocio",
    "nombre de la empresa": "negocio", "nombre comercial": "negocio",
    "name": "negocio", "business name": "negocio",
    "nombre de contacto": "contacto", "contacto": "contacto", "nombre del contacto": "contacto",
    "numero de contacto": "telefono", "telefono": "telefono", "celular": "telefono",
    "numero": "telefono", "whatsapp": "telefono", "phone": "telefono", "phone number": "telefono",
    "correo electronico": "correo", "correo": "correo", "email": "correo", "e-mail": "correo",
    "direccion": "direccion", "address": "direccion",
    "ciudad": "ciudad", "city": "ciudad",
    "estado": "estado",
    "notas": "notas", "observaciones": "notas", "comentarios": "notas",
    "pagina web": "pagina_web", "web": "pagina_web", "sitio web": "pagina_web", "website": "pagina_web",
    "cargo de contacto": "cargo_contacto", "cargo": "cargo_contacto", "puesto": "cargo_contacto",
    "ruc": "ruc", "nit": "ruc", "cif": "ruc", "ruc/nit/cif": "ruc",
    "etiquetas": "tags", "tags": "tags",
    "favorito": "favorito",
    "ultimo contacto": "ultimo_contacto",
    "proximo seguimiento": "proximo_contacto", "proxima fecha de seguimiento": "proximo_contacto",
    "fecha de registro": "fecha", "fecha": "fecha",
}

# Columnas del CSV exportado por nuestra propia app (Google Maps / directorios)
# que deliberadamente NO importamos porque no aportan a esta base de datos:
# Instagram, Facebook, Twitter, Linkedin, Yelp, Youtube, PlaceID, CID,
# ReviewCount, AverageRating, Latitude, Longitude. Al no tener alias en
# HEADER_ALIASES, normalize_header() las ignora automáticamente.


def today_str():
    return datetime.date.today().strftime("%d/%m/%Y")


def sanitize_prefix(text):
    return "".join(ch for ch in text.upper() if ch.isalnum())[:8]


def pad_number(n):
    return f"{n:02d}" if n < 100 else str(n)


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn")


def normalize_header(h):
    return HEADER_ALIASES.get(strip_accents(h).strip().lower())


def parse_ddmmyyyy(s):
    try:
        d, m, y = str(s).split("/")
        return datetime.date(int(y), int(m), int(d))
    except Exception:
        return None


def clean_phone_digits(phone):
    return "".join(ch for ch in str(phone) if ch.isdigit())


def ensure_url(url):
    url = (url or "").strip()
    if not url:
        return ""
    if not url.startswith("http://") and not url.startswith("https://"):
        url = "https://" + url
    return url


# ---------------------------------------------------------------------------
# Capa de datos (SQLite)
# ---------------------------------------------------------------------------

NEW_COLUMNS = {
    "pagina_web": "TEXT DEFAULT ''",
    "cargo_contacto": "TEXT DEFAULT ''",
    "ruc": "TEXT DEFAULT ''",
    "favorito": "INTEGER NOT NULL DEFAULT 0",
    "tags": "TEXT DEFAULT ''",
    "ultimo_contacto": "TEXT DEFAULT ''",
    "proximo_contacto": "TEXT DEFAULT ''",
}


class Database:
    def __init__(self, path):
        self.path = path
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA journal_mode = WAL")
        self.conn.execute("PRAGMA synchronous = NORMAL")
        self.conn.execute("PRAGMA temp_store = MEMORY")
        self.conn.execute("PRAGMA cache_size = -8000")
        self.conn.execute("PRAGMA foreign_keys = ON")
        self._create_tables()
        self._migrate()
        self._seed_if_empty()

    def _create_tables(self):
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                prefix TEXT NOT NULL,
                counter INTEGER NOT NULL DEFAULT 1
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS businesses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE NOT NULL,
                category_id INTEGER NOT NULL,
                negocio TEXT NOT NULL,
                contacto TEXT,
                telefono TEXT NOT NULL,
                correo TEXT NOT NULL,
                direccion TEXT NOT NULL,
                ciudad TEXT,
                estado TEXT NOT NULL DEFAULT 'Activo',
                notas TEXT,
                fecha TEXT,
                FOREIGN KEY(category_id) REFERENCES categories(id) ON DELETE CASCADE
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                negocio TEXT,
                action TEXT NOT NULL,
                detail TEXT,
                timestamp TEXT NOT NULL
            )
        """)
        self.conn.execute("CREATE INDEX IF NOT EXISTS idx_businesses_category ON businesses(category_id)")
        self.conn.execute("CREATE INDEX IF NOT EXISTS idx_businesses_code ON businesses(code)")
        self.conn.execute("CREATE INDEX IF NOT EXISTS idx_history_code ON history(code)")
        self.conn.commit()

    def _migrate(self):
        """Agrega columnas nuevas a bases de datos creadas con versiones anteriores,
        sin tocar ni perder los datos que ya existan."""
        cols = {row["name"] for row in self.conn.execute("PRAGMA table_info(businesses)")}
        for col, coltype in NEW_COLUMNS.items():
            if col not in cols:
                self.conn.execute(f"ALTER TABLE businesses ADD COLUMN {col} {coltype}")
        self.conn.commit()

    def _seed_if_empty(self):
        cur = self.conn.execute("SELECT COUNT(*) FROM categories")
        if cur.fetchone()[0] == 0:
            self.conn.execute(
                "INSERT INTO categories (name, prefix, counter) VALUES (?,?,?)",
                ("Cafetería", "TSCA", 2),
            )
            cat_id = self.conn.execute(
                "SELECT id FROM categories WHERE name = ?", ("Cafetería",)
            ).fetchone()[0]
            demo = {
                "negocio": "The Standard Cafetería", "contacto": "María Gómez",
                "telefono": "+51 987 654 321", "correo": "maria.gomez@thestandard.com",
                "direccion": "Av. Los Álamos 123, San Isidro", "ciudad": "Lima", "estado": "Activo",
                "notas": "Registro de ejemplo — puedes eliminarlo.",
                "pagina_web": "https://thestandardcafe.com", "cargo_contacto": "Gerente general",
                "ruc": "", "tags": "demo", "ultimo_contacto": today_str(), "proximo_contacto": "",
            }
            self.insert_business("TSCA01", cat_id, demo, log=False)
            self.conn.commit()

    # ---- categorías ----
    def get_categories(self):
        rows = self.conn.execute("SELECT id, name, prefix, counter FROM categories ORDER BY name").fetchall()
        return {name: {"id": cid, "prefix": prefix, "counter": counter} for cid, name, prefix, counter in rows}

    def create_category(self, name, prefix):
        cur = self.conn.execute(
            "INSERT INTO categories (name, prefix, counter) VALUES (?,?,1)", (name, prefix)
        )
        self.conn.commit()
        return cur.lastrowid

    def delete_category(self, cat_id):
        self.conn.execute("DELETE FROM businesses WHERE category_id = ?", (cat_id,))
        self.conn.execute("DELETE FROM categories WHERE id = ?", (cat_id,))
        self.conn.commit()

    def bump_counter(self, cat_id, new_counter):
        self.conn.execute("UPDATE categories SET counter = ? WHERE id = ?", (new_counter, cat_id))
        self.conn.commit()

    def next_code_for(self, cat_meta):
        return cat_meta["prefix"] + pad_number(cat_meta["counter"])

    def code_exists(self, code):
        return self.conn.execute("SELECT 1 FROM businesses WHERE code = ?", (code,)).fetchone() is not None

    # ---- negocios ----
    SELECT_COLUMNS = """
        b.code, c.name as category, b.negocio, b.contacto, b.telefono, b.correo,
        b.direccion, b.ciudad, b.estado, b.notas, b.fecha, b.pagina_web, b.cargo_contacto,
        b.ruc, b.favorito, b.tags, b.ultimo_contacto, b.proximo_contacto
    """

    def get_businesses(self):
        return self.conn.execute(f"""
            SELECT {self.SELECT_COLUMNS}
            FROM businesses b JOIN categories c ON c.id = b.category_id
            ORDER BY c.name, b.code
        """).fetchall()

    def get_businesses_grouped(self):
        rows = self.get_businesses()
        grouped = {}
        for row in rows:
            grouped.setdefault(row["category"], []).append(row)
        return rows, grouped

    def get_business(self, code):
        return self.conn.execute(f"""
            SELECT {self.SELECT_COLUMNS}
            FROM businesses b JOIN categories c ON c.id = b.category_id
            WHERE b.code = ?
        """, (code,)).fetchone()

    def insert_business(self, code, category_id, data, log=True):
        self.conn.execute("""
            INSERT INTO businesses
            (code, category_id, negocio, contacto, telefono, correo, direccion, ciudad, estado,
             notas, fecha, pagina_web, cargo_contacto, ruc, favorito, tags, ultimo_contacto, proximo_contacto)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (code, category_id, data.get("negocio", ""), data.get("contacto", ""),
              data.get("telefono", ""), data.get("correo", ""), data.get("direccion", ""),
              data.get("ciudad", ""), data.get("estado", "Activo") or "Activo",
              data.get("notas", ""), today_str(), data.get("pagina_web", ""),
              data.get("cargo_contacto", ""), data.get("ruc", ""), int(bool(data.get("favorito", 0))),
              data.get("tags", ""), data.get("ultimo_contacto", ""), data.get("proximo_contacto", "")))
        self.conn.commit()
        if log:
            self.log_history(code, data.get("negocio", ""), "Creado", "Registro nuevo.")

    def update_business(self, code, data, log=True):
        before = self.get_business(code)
        self.conn.execute("""
            UPDATE businesses SET negocio=?, contacto=?, telefono=?, correo=?, direccion=?, ciudad=?,
                   estado=?, notas=?, pagina_web=?, cargo_contacto=?, ruc=?, tags=?,
                   ultimo_contacto=?, proximo_contacto=?
            WHERE code = ?
        """, (data.get("negocio", ""), data.get("contacto", ""), data.get("telefono", ""),
              data.get("correo", ""), data.get("direccion", ""), data.get("ciudad", ""),
              data.get("estado", "Activo") or "Activo", data.get("notas", ""), data.get("pagina_web", ""),
              data.get("cargo_contacto", ""), data.get("ruc", ""), data.get("tags", ""),
              data.get("ultimo_contacto", ""), data.get("proximo_contacto", ""), code))
        self.conn.commit()
        if log and before:
            detail = self._diff_detail(before, data)
            if detail:
                self.log_history(code, data.get("negocio", before["negocio"]), "Editado", detail)

    @staticmethod
    def _diff_detail(before, after):
        changes = []
        for key in FULL_FIELDS:
            old_val = (before[key] or "") if key in before.keys() else ""
            new_val = after.get(key, old_val) or ""
            if str(old_val) != str(new_val):
                label = FIELD_LABELS.get(key, key)
                changes.append(f"{label}: \"{old_val}\" → \"{new_val}\"")
        return "; ".join(changes)

    def set_favorite(self, code, value):
        self.conn.execute("UPDATE businesses SET favorito = ? WHERE code = ?", (int(bool(value)), code))
        self.conn.commit()

    def set_tags(self, code, tags_value):
        self.conn.execute("UPDATE businesses SET tags = ? WHERE code = ?", (tags_value, code))
        self.conn.commit()

    def bulk_update(self, codes, changes):
        if not codes or not changes:
            return
        set_clause = ", ".join(f"{k} = ?" for k in changes)
        values = list(changes.values())
        for code in codes:
            self.conn.execute(f"UPDATE businesses SET {set_clause} WHERE code = ?", values + [code])
        self.conn.commit()

    def delete_business(self, code, log=True):
        row = self.get_business(code)
        if row and log:
            self.log_history(code, row["negocio"], "Eliminado", "Registro eliminado.")
        self.conn.execute("DELETE FROM businesses WHERE code = ?", (code,))
        self.conn.commit()

    def find_by_name(self, name, exclude_code=None):
        """Busca un negocio existente con el mismo nombre (sin importar mayúsculas
        ni espacios). Se usa para evitar duplicados al importar o para avisar
        de posibles duplicados al registrar/editar manualmente."""
        name_norm = (name or "").strip().lower()
        if not name_norm:
            return None
        row = self.conn.execute(f"""
            SELECT {self.SELECT_COLUMNS}
            FROM businesses b JOIN categories c ON c.id = b.category_id
            WHERE LOWER(TRIM(b.negocio)) = ?
            ORDER BY b.code LIMIT 1
        """, (name_norm,)).fetchone()
        if row and exclude_code and row["code"] == exclude_code:
            return None
        return row

    def duplicate_name_groups(self):
        """Devuelve grupos de negocios que comparten el mismo nombre (posibles
        duplicados ya existentes en la base), para el panel de notificaciones."""
        rows = self.get_businesses()
        groups = {}
        for row in rows:
            key = (row["negocio"] or "").strip().lower()
            if key:
                groups.setdefault(key, []).append(row)
        return [g for g in groups.values() if len(g) > 1]

    def move_category(self, code, new_category_id):
        self.conn.execute("UPDATE businesses SET category_id = ? WHERE code = ?", (new_category_id, code))
        self.conn.commit()

    def distinct_cities(self):
        rows = self.conn.execute(
            "SELECT DISTINCT ciudad FROM businesses WHERE ciudad IS NOT NULL AND ciudad != '' ORDER BY ciudad"
        ).fetchall()
        return [r["ciudad"] for r in rows]

    # ---- historial ----
    def log_history(self, code, negocio, action, detail=""):
        self.conn.execute(
            "INSERT INTO history (code, negocio, action, detail, timestamp) VALUES (?,?,?,?,?)",
            (code, negocio, action, detail, datetime.datetime.now().strftime("%d/%m/%Y %H:%M"))
        )
        self.conn.commit()

    def get_history(self, code=None, limit=200):
        if code:
            return self.conn.execute(
                "SELECT * FROM history WHERE code = ? ORDER BY id DESC LIMIT ?", (code, limit)
            ).fetchall()
        return self.conn.execute(
            "SELECT * FROM history ORDER BY id DESC LIMIT ?", (limit,)
        ).fetchall()


# ---------------------------------------------------------------------------
# Paleta — tema claro / oscuro conmutable
# ---------------------------------------------------------------------------

LIGHT_THEME = {
    "BG": "#FFFFFF", "SURFACE": "#FFFFFF", "SURFACE_ALT": "#F7F8FA", "BORDER": "#E4E7EC",
    "TEXT": "#101828", "TEXT_SOFT": "#667085", "TEXT_FAINT": "#98A2B3",
    "ACCENT": "#2563EB", "ACCENT_HOVER": "#1D4ED8", "ACCENT_SOFT": "#EFF4FF",
    "SUCCESS": "#12805C", "SUCCESS_SOFT": "#E7F6EF",
    "DANGER": "#B42318", "DANGER_SOFT": "#FEF3F2",
    "WARN_SOFT": "#FFFAEB", "WARN_TEXT": "#B54708", "FAV_COLOR": "#B54708",
}

DARK_THEME = {
    "BG": "#0B0F17", "SURFACE": "#141B26", "SURFACE_ALT": "#1B2431", "BORDER": "#2B3648",
    "TEXT": "#E7ECF3", "TEXT_SOFT": "#9AA6B8", "TEXT_FAINT": "#6C7789",
    "ACCENT": "#4C8DFF", "ACCENT_HOVER": "#3E7BE8", "ACCENT_SOFT": "#1E2A40",
    "SUCCESS": "#34C787", "SUCCESS_SOFT": "#123326",
    "DANGER": "#F0665E", "DANGER_SOFT": "#3A1B1B",
    "WARN_SOFT": "#332A12", "WARN_TEXT": "#F2B84B", "FAV_COLOR": "#F2B84B",
}

THEMES = {"light": LIGHT_THEME, "dark": DARK_THEME}

# Estos nombres quedan como variables de módulo "mutables": set_theme() las
# reasigna, y como toda la interfaz se reconstruye al cambiar de tema, los
# widgets nuevos siempre leen el valor vigente en el momento de crearse.
BG = SURFACE = SURFACE_ALT = BORDER = TEXT = TEXT_SOFT = TEXT_FAINT = None
ACCENT = ACCENT_HOVER = ACCENT_SOFT = None
SUCCESS = SUCCESS_SOFT = DANGER = DANGER_SOFT = WARN_SOFT = WARN_TEXT = FAV_COLOR = None
ESTADO_COLORS = {}


def set_theme(name):
    """Aplica la paleta 'light' o 'dark' a las variables de módulo."""
    palette = THEMES.get(name, LIGHT_THEME)
    globals().update(palette)
    global ESTADO_COLORS
    ESTADO_COLORS = {"Activo": palette["SUCCESS"], "Potencial": palette["WARN_TEXT"],
                      "Inactivo": palette["TEXT_FAINT"]}


set_theme("light")

FONT_BASE   = ("Segoe UI", 10)
FONT_SMALL  = ("Segoe UI", 9)
FONT_LABEL  = ("Segoe UI", 9, "bold")
FONT_TITLE  = ("Segoe UI", 18, "bold")
FONT_SUB    = ("Segoe UI", 12, "bold")
FONT_MONO   = ("Consolas", 13, "bold")


# ---------------------------------------------------------------------------
# Configuración persistente (tema, tamaño/posición de ventana)
# ---------------------------------------------------------------------------

CONFIG_PATH = os.path.join(APP_DIR, "config.json")


def load_config():
    import json
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(cfg):
    import json
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Widgets reutilizables
# ---------------------------------------------------------------------------

class ModernButton(tk.Button):
    """Botón plano con estado hover, pensado para el look moderno blanco."""

    def __init__(self, master, bg=ACCENT, fg="white", hover=None, **kwargs):
        self._bg = bg
        self._hover = hover or self._darken(bg)
        self._fg = fg
        kwargs.setdefault("highlightthickness", 0)
        super().__init__(
            master, bg=bg, fg=fg, activebackground=self._hover, activeforeground=fg,
            relief="flat", bd=0, cursor="hand2", font=kwargs.pop("font", FONT_LABEL),
            padx=kwargs.pop("padx", 14), pady=kwargs.pop("pady", 8),
            **kwargs,
        )
        self.bind("<Enter>", lambda e: self.configure(bg=self._hover))
        self.bind("<Leave>", lambda e: self.configure(bg=self._bg))

    @staticmethod
    def _darken(hex_color, factor=0.88):
        hex_color = hex_color.lstrip("#")
        if len(hex_color) != 6:
            return hex_color
        r, g, b = (int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
        r, g, b = (max(0, int(c * factor)) for c in (r, g, b))
        return f"#{r:02x}{g:02x}{b:02x}"


class ScrollableFrame(tk.Frame):
    """Frame con scroll vertical (rueda del mouse incluida)."""

    def __init__(self, master, bg=SURFACE, **kwargs):
        super().__init__(master, bg=bg, **kwargs)

        self.canvas = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = tk.Frame(self.canvas, bg=bg)

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self._window = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vsb.grid(row=0, column=1, sticky="ns")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        for widget in (self.canvas, self.inner):
            widget.bind("<Enter>", lambda e: self._bind_mousewheel())
            widget.bind("<Leave>", lambda e: self._unbind_mousewheel())

    def _on_inner_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self._window, width=event.width)

    def _bind_mousewheel(self):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)

    def _unbind_mousewheel(self):
        self.canvas.unbind_all("<MouseWheel>")
        self.canvas.unbind_all("<Button-4>")
        self.canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_mousewheel_linux(self, event):
        self.canvas.yview_scroll(-1 if event.num == 4 else 1, "units")


class Card(tk.Frame):
    """Contenedor con look de 'tarjeta' blanca moderna."""

    def __init__(self, master, **kwargs):
        super().__init__(master, bg=SURFACE, highlightbackground=BORDER,
                          highlightthickness=1, bd=0, **kwargs)


def styled_entry(parent, **kwargs):
    e = tk.Entry(parent, relief="flat", bd=0, font=FONT_BASE, bg=SURFACE_ALT, fg=TEXT,
                 insertbackground=TEXT, highlightthickness=1,
                 highlightbackground=BORDER, highlightcolor=ACCENT, **kwargs)

    def on_focus_in(_e):
        e.configure(highlightbackground=ACCENT, highlightcolor=ACCENT, bg=SURFACE)

    def on_focus_out(_e):
        e.configure(highlightbackground=BORDER, bg=SURFACE_ALT)

    e.bind("<FocusIn>", on_focus_in)
    e.bind("<FocusOut>", on_focus_out)
    return e


class MiniCalendar(tk.Toplevel):
    """Selector de fecha simple, sin dependencias externas."""

    def __init__(self, master, on_select, initial=None):
        super().__init__(master)
        self.overrideredirect(True)
        self.configure(bg=SURFACE, highlightbackground=BORDER, highlightthickness=1)
        self.attributes("-topmost", True)
        self.on_select = on_select

        today = parse_ddmmyyyy(initial) or datetime.date.today()
        self.year = today.year
        self.month = today.month

        try:
            x = master.winfo_rootx()
            y = master.winfo_rooty() + master.winfo_height() + 2
            self.geometry(f"+{x}+{y}")
        except Exception:
            pass

        self._build()
        self.bind("<FocusOut>", lambda e: self.destroy())
        self.after(30, lambda: self.focus_set())

    def _build(self):
        for w in self.winfo_children():
            w.destroy()

        header = tk.Frame(self, bg=SURFACE)
        header.pack(fill="x", padx=6, pady=(6, 2))
        tk.Button(header, text="‹", relief="flat", bd=0, bg=SURFACE, fg=TEXT, cursor="hand2",
                  command=self._prev_month).pack(side="left")
        tk.Label(header, text=f"{calendar_module.month_name[self.month].capitalize()} {self.year}",
                 bg=SURFACE, fg=TEXT, font=FONT_LABEL, width=16).pack(side="left")
        tk.Button(header, text="›", relief="flat", bd=0, bg=SURFACE, fg=TEXT, cursor="hand2",
                  command=self._next_month).pack(side="right")

        grid = tk.Frame(self, bg=SURFACE)
        grid.pack(padx=6, pady=(0, 6))
        for i, d in enumerate(["L", "M", "X", "J", "V", "S", "D"]):
            tk.Label(grid, text=d, bg=SURFACE, fg=TEXT_FAINT, font=("Segoe UI", 8, "bold"), width=3)\
                .grid(row=0, column=i)

        cal = calendar_module.Calendar(firstweekday=0)
        row = 1
        for week in cal.monthdayscalendar(self.year, self.month):
            for col, day in enumerate(week):
                if day == 0:
                    tk.Label(grid, text="", bg=SURFACE, width=3).grid(row=row, column=col)
                else:
                    tk.Button(grid, text=str(day), relief="flat", bd=0, width=3, bg=SURFACE, fg=TEXT,
                              activebackground=ACCENT_SOFT, cursor="hand2",
                              command=lambda d=day: self._pick(d)).grid(row=row, column=col, pady=1)
            row += 1

        tk.Button(self, text="Hoy", relief="flat", bd=0, bg=ACCENT_SOFT, fg=ACCENT, font=FONT_LABEL,
                  cursor="hand2", command=self._pick_today).pack(fill="x", padx=6, pady=(0, 6))

    def _prev_month(self):
        self.month -= 1
        if self.month == 0:
            self.month, self.year = 12, self.year - 1
        self._build()

    def _next_month(self):
        self.month += 1
        if self.month == 13:
            self.month, self.year = 1, self.year + 1
        self._build()

    def _pick(self, day):
        self.on_select(datetime.date(self.year, self.month, day).strftime("%d/%m/%Y"))
        self.destroy()

    def _pick_today(self):
        self.on_select(datetime.date.today().strftime("%d/%m/%Y"))
        self.destroy()


def build_date_field(parent, label_text, initial=""):
    """Crea una etiqueta + entry + botones [📅][Hoy] para capturar una fecha."""
    tk.Label(parent, text=label_text, font=FONT_LABEL, bg=SURFACE, fg=TEXT_SOFT)\
        .pack(anchor="w", pady=(8, 2))
    row = tk.Frame(parent, bg=SURFACE)
    row.pack(fill="x", pady=(0, 2))
    entry = styled_entry(row)
    entry.insert(0, initial or "")
    entry.pack(side="left", fill="x", expand=True, ipady=4)

    def open_cal():
        def on_pick(value):
            entry.delete(0, "end")
            entry.insert(0, value)
        MiniCalendar(row, on_pick, entry.get().strip() or None)

    def set_today():
        entry.delete(0, "end")
        entry.insert(0, today_str())

    ModernButton(row, text="📅", bg=SURFACE, fg=TEXT_SOFT, hover=SURFACE_ALT,
                 highlightbackground=BORDER, highlightthickness=1, padx=8, pady=4,
                 command=open_cal).pack(side="left", padx=(6, 0))
    ModernButton(row, text="Hoy", bg=ACCENT_SOFT, fg=ACCENT, hover="#E0EAFE", padx=8, pady=4,
                 command=set_today).pack(side="left", padx=(6, 0))
    return entry


# ---------------------------------------------------------------------------
# Interfaz gráfica
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.config_data = load_config()
        self._theme_name = self.config_data.get("theme", "light")
        set_theme(self._theme_name)

        self.title("Registro de Negocios — Base de datos local")
        self.geometry(self.config_data.get("geometry") or "1360x800")
        self.configure(bg=SURFACE_ALT)
        self.minsize(980, 600)

        self.db = Database(DB_PATH)
        self.categories = {}
        self.tab_widgets = {}       # nombre_rubro -> {"frame", "tree"}
        self.field_order = []
        self.dashboard_frame = None
        self._context_menu_codes = []
        self.notifications = []     # lista de avisos de la sesión (duplicados, etc.)

        self._build_style()
        self._build_layout()
        self._build_shortcuts()
        self.refresh_all(rebuild_tabs=True)

        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.after(80, lambda: self.rubro_combo.focus_set())

    def on_close(self):
        try:
            self.config_data["geometry"] = self.geometry()
            self.config_data["theme"] = self.current_theme()
            save_config(self.config_data)
        except Exception:
            pass
        self.destroy()

    def current_theme(self):
        return getattr(self, "_theme_name", self.config_data.get("theme", "light"))

    # ================================================================
    # Tema claro / oscuro
    # ================================================================
    def toggle_theme(self):
        new_theme = "dark" if self.current_theme() == "light" else "light"
        self._theme_name = new_theme
        set_theme(new_theme)
        self.config_data["theme"] = new_theme
        save_config(self.config_data)
        self._rebuild_ui()

    def _rebuild_ui(self):
        """Reconstruye toda la interfaz con la paleta vigente (usado al
        cambiar de tema, para que absolutamente todos los widgets —
        incluidos los ya creados— respeten los nuevos colores)."""
        search_text = self.search_var.get() if hasattr(self, "search_var") else ""
        for w in self.winfo_children():
            w.destroy()

        self.configure(bg=SURFACE_ALT)
        self.tab_widgets = {}
        self.field_order = []
        self.dashboard_frame = None

        self._build_style()
        self._build_layout()
        self._build_shortcuts()
        if search_text:
            self.search_var.set(search_text)
        self.refresh_all(rebuild_tabs=True)

    # ================================================================
    # Estilo ttk
    # ================================================================
    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure("TPanedwindow", background=SURFACE_ALT)
        style.configure("TNotebook", background=SURFACE_ALT, borderwidth=0, tabmargins=(0, 6, 0, 0))
        style.configure("TNotebook.Tab", padding=(14, 9), font=FONT_SMALL,
                         background=SURFACE_ALT, foreground=TEXT_SOFT, borderwidth=0)
        style.map("TNotebook.Tab",
                  background=[("selected", SURFACE)],
                  foreground=[("selected", ACCENT)])

        style.configure("Treeview", rowheight=28, font=FONT_SMALL, background=SURFACE,
                         fieldbackground=SURFACE, foreground=TEXT, borderwidth=0)
        style.configure("Treeview.Heading", font=FONT_LABEL, background=SURFACE_ALT,
                         foreground=TEXT_SOFT, relief="flat", borderwidth=0)
        style.map("Treeview",
                  background=[("selected", ACCENT_SOFT)],
                  foreground=[("selected", ACCENT)])
        style.map("Treeview.Heading", background=[("active", SURFACE_ALT)])

        style.configure("TCombobox", fieldbackground=SURFACE_ALT, background=SURFACE,
                         foreground=TEXT, arrowcolor=TEXT_SOFT, borderwidth=0)
        style.configure("Vertical.TScrollbar", background=SURFACE_ALT, troughcolor=SURFACE,
                         bordercolor=SURFACE, arrowcolor=TEXT_FAINT)
        style.configure("Horizontal.TScrollbar", background=SURFACE_ALT, troughcolor=SURFACE,
                         bordercolor=SURFACE, arrowcolor=TEXT_FAINT)

    # ================================================================
    # Layout general
    # ================================================================
    def _build_layout(self):
        header = tk.Frame(self, bg=SURFACE_ALT)
        header.pack(fill="x", padx=20, pady=(16, 4))

        title_box = tk.Frame(header, bg=SURFACE_ALT)
        title_box.pack(side="left", anchor="w")
        tk.Label(title_box, text="Registro de Negocios", font=FONT_TITLE, bg=SURFACE_ALT, fg=TEXT)\
            .pack(anchor="w")
        tk.Label(title_box, text=f"Base de datos local: {DB_PATH}", font=("Segoe UI", 8),
                 bg=SURFACE_ALT, fg=TEXT_FAINT).pack(anchor="w")

        header_actions = tk.Frame(header, bg=SURFACE_ALT)
        header_actions.pack(side="right", anchor="e")

        theme_label = "☀️ Modo claro" if self.current_theme() == "dark" else "🌙 Modo oscuro"
        ModernButton(header_actions, text=theme_label, bg=SURFACE, fg=TEXT, hover=SURFACE_ALT,
                     highlightbackground=BORDER, highlightthickness=1, font=FONT_SMALL,
                     padx=10, pady=6, command=self.toggle_theme).pack(side="right", padx=(8, 0))

        self.notif_btn = ModernButton(header_actions, text="🔔 Notificaciones", bg=SURFACE, fg=TEXT,
                                       hover=SURFACE_ALT, highlightbackground=BORDER, highlightthickness=1,
                                       font=FONT_SMALL, padx=10, pady=6, command=self.show_notifications_dialog)
        self.notif_btn.pack(side="right", padx=(8, 0))

        stats_box = tk.Frame(header_actions, bg=SURFACE_ALT)
        stats_box.pack(side="right", padx=(0, 8))
        self.stats_var = tk.StringVar()
        tk.Label(stats_box, textvariable=self.stats_var, font=FONT_SMALL, bg=SURFACE_ALT,
                 fg=TEXT_SOFT).pack(anchor="e")

        self.paned = ttk.PanedWindow(self, orient="horizontal")
        self.paned.pack(fill="both", expand=True, padx=20, pady=(10, 8))

        form_holder = tk.Frame(self.paned, bg=SURFACE_ALT)
        dir_holder = tk.Frame(self.paned, bg=SURFACE_ALT)
        self.paned.add(form_holder, weight=0)
        self.paned.add(dir_holder, weight=1)

        self._build_form(form_holder)
        self._build_directory(dir_holder)

        footer = tk.Frame(self, bg=SURFACE_ALT)
        footer.pack(fill="x", padx=20, pady=(0, 12))
        shortcuts_text = (
            "⌨  Ctrl+Enter registrar   ·   Ctrl+N nuevo   ·   Ctrl+E editar   ·   Ctrl+D duplicar   ·   "
            "Supr eliminar   ·   Ctrl+F buscar   ·   Ctrl+S exportar   ·   F5 actualizar"
        )
        tk.Label(footer, text=shortcuts_text, font=("Segoe UI", 8), bg=SURFACE_ALT, fg=TEXT_FAINT)\
            .pack(anchor="w")

    # ================================================================
    # Atajos de teclado
    # ================================================================
    def _build_shortcuts(self):
        self.bind_all("<Control-Return>", lambda e: self.submit_business())
        self.bind_all("<Control-n>", lambda e: self.focus_new_entry())
        self.bind_all("<Control-N>", lambda e: self.focus_new_entry())
        self.bind_all("<Control-e>", lambda e: self.edit_selected())
        self.bind_all("<Control-E>", lambda e: self.edit_selected())
        self.bind_all("<Control-d>", lambda e: self.duplicate_selected())
        self.bind_all("<Control-D>", lambda e: self.duplicate_selected())
        self.bind_all("<Control-s>", lambda e: self.export_excel())
        self.bind_all("<Control-S>", lambda e: self.export_excel())
        self.bind_all("<Control-f>", lambda e: self._focus_search())
        self.bind_all("<Control-F>", lambda e: self._focus_search())
        self.bind_all("<F5>", lambda e: self.refresh_all(rebuild_tabs=True))
        self.bind_all("<Escape>", self._on_escape)
        self.bind_all("<Delete>", self._on_delete_key)

    def _focus_search(self):
        if hasattr(self, "search_entry"):
            self.search_entry.focus_set()

    def _on_escape(self, event):
        widget = self.focus_get()
        self.msg_var.set("")
        if isinstance(widget, (tk.Entry, tk.Text)):
            return
        self.focus_set()

    def _on_delete_key(self, event):
        widget = self.focus_get()
        if isinstance(widget, ttk.Treeview):
            self.delete_selected()

    def focus_new_entry(self):
        self.rubro_combo.focus_set()

    # ================================================================
    # Formulario (izquierda) — con scroll
    # ================================================================
    def _build_form(self, parent):
        card = Card(parent)
        card.pack(fill="both", expand=True)

        header = tk.Frame(card, bg=SURFACE)
        header.pack(fill="x", padx=20, pady=(18, 6))
        tk.Label(header, text="Ficha de registro", font=FONT_SUB, bg=SURFACE, fg=TEXT).pack(anchor="w")
        tk.Label(header, text="Los campos con * son obligatorios.", font=("Segoe UI", 8),
                 bg=SURFACE, fg=TEXT_FAINT).pack(anchor="w", pady=(2, 0))

        scroll = ScrollableFrame(card, bg=SURFACE)
        scroll.pack(fill="both", expand=True)
        inner = scroll.inner
        inner.configure(padx=20, pady=4)

        # Rubro
        tk.Label(inner, text="Rubro *", font=FONT_LABEL, bg=SURFACE, fg=TEXT_SOFT).pack(anchor="w", pady=(8, 2))
        self.rubro_var = tk.StringVar()
        self.rubro_combo = ttk.Combobox(inner, textvariable=self.rubro_var, state="readonly", font=FONT_BASE)
        self.rubro_combo.pack(fill="x", pady=(0, 4), ipady=3)
        self.rubro_combo.bind("<<ComboboxSelected>>", self.on_rubro_change)

        self.new_cat_frame = tk.Frame(inner, bg=WARN_SOFT, highlightbackground="#F2C94C",
                                       highlightthickness=1, bd=0)
        tk.Label(self.new_cat_frame, text="Rubro nuevo: indica el nombre y las siglas para su código\n"
                                          "(ej. Cafetería → TSCA: TS de The Standard + CA del rubro).",
                 bg=WARN_SOFT, fg=WARN_TEXT, font=("Segoe UI", 8), justify="left")\
            .pack(anchor="w", padx=10, pady=(8, 6))
        tk.Label(self.new_cat_frame, text="Nombre del rubro", bg=WARN_SOFT, fg=WARN_TEXT,
                 font=FONT_LABEL).pack(anchor="w", padx=10)
        self.new_cat_name = styled_entry(self.new_cat_frame)
        self.new_cat_name.pack(fill="x", padx=10, pady=(2, 8), ipady=3)
        tk.Label(self.new_cat_frame, text="Siglas para el código", bg=WARN_SOFT, fg=WARN_TEXT,
                 font=FONT_LABEL).pack(anchor="w", padx=10)
        self.new_cat_prefix = styled_entry(self.new_cat_frame)
        self.new_cat_prefix.pack(fill="x", padx=10, pady=(2, 10), ipady=3)
        self.new_cat_prefix.bind("<KeyRelease>", lambda e: self.update_code_preview())
        self.new_cat_name.bind("<KeyRelease>", lambda e: self.update_code_preview())

        self.fields = {}
        field_defs = [
            ("negocio", "Nombre del negocio *"),
            ("contacto", "Nombre de contacto"),
            ("telefono", "Número de contacto *"),
            ("correo", "Correo electrónico *"),
            ("direccion", "Dirección *"),
            ("ciudad", "Ciudad"),
            ("pagina_web", "Página web"),
            ("cargo_contacto", "Cargo de contacto"),
            ("ruc", "RUC / NIT / CIF"),
            ("tags", "Etiquetas (separadas por coma)"),
        ]
        for key, label in field_defs:
            tk.Label(inner, text=label, font=FONT_LABEL, bg=SURFACE, fg=TEXT_SOFT).pack(anchor="w", pady=(8, 2))
            entry = styled_entry(inner)
            entry.pack(fill="x", pady=(0, 2), ipady=4)
            self.fields[key] = entry
            self.field_order.append(entry)

        tk.Label(inner, text="Estado del cliente", font=FONT_LABEL, bg=SURFACE, fg=TEXT_SOFT)\
            .pack(anchor="w", pady=(8, 2))
        self.estado_var = tk.StringVar(value="Activo")
        ttk.Combobox(inner, textvariable=self.estado_var, values=ESTADOS, state="readonly", font=FONT_BASE)\
            .pack(fill="x", pady=(0, 2), ipady=3)

        self.ultimo_contacto_entry = build_date_field(inner, "Último contacto")
        self.proximo_contacto_entry = build_date_field(inner, "Próxima fecha de seguimiento")

        self.favorito_var = tk.BooleanVar(value=False)
        tk.Checkbutton(inner, text="⭐ Marcar como favorito", variable=self.favorito_var, bg=SURFACE,
                        fg=TEXT_SOFT, selectcolor=SURFACE, activebackground=SURFACE, font=FONT_SMALL)\
            .pack(anchor="w", pady=(10, 0))

        tk.Label(inner, text="Notas", font=FONT_LABEL, bg=SURFACE, fg=TEXT_SOFT).pack(anchor="w", pady=(8, 2))
        self.notas_text = tk.Text(inner, height=3, relief="flat", bd=0, font=FONT_SMALL,
                                   bg=SURFACE_ALT, fg=TEXT, highlightthickness=1,
                                   highlightbackground=BORDER, highlightcolor=ACCENT, wrap="word")
        self.notas_text.pack(fill="x", pady=(0, 4))

        self._wire_enter_navigation()

        preview_frame = tk.Frame(inner, bg=SURFACE_ALT, highlightbackground=BORDER, highlightthickness=1)
        preview_frame.pack(fill="x", pady=(14, 10))
        tk.Label(preview_frame, text="PRÓXIMO CÓDIGO", font=("Segoe UI", 7, "bold"), bg=SURFACE_ALT,
                 fg=TEXT_FAINT).pack(anchor="w", padx=12, pady=(8, 0))
        self.code_preview_var = tk.StringVar(value="—")
        tk.Label(preview_frame, textvariable=self.code_preview_var, font=FONT_MONO,
                 bg=ACCENT_SOFT, fg=ACCENT, padx=12, pady=5).pack(anchor="w", padx=12, pady=(2, 10))

        self.submit_btn = ModernButton(inner, text="Registrar negocio  (Ctrl+Enter)",
                                        bg=ACCENT, fg="white", command=self.submit_business)
        self.submit_btn.pack(fill="x", pady=(2, 8))

        self.msg_var = tk.StringVar()
        self.msg_label = tk.Label(inner, textvariable=self.msg_var, font=FONT_SMALL,
                                   bg=SURFACE, fg=SUCCESS, wraplength=300, justify="left")
        self.msg_label.pack(anchor="w", pady=(0, 16))

    def _wire_enter_navigation(self):
        ordered = self.field_order
        for i, entry in enumerate(ordered):
            nxt = ordered[i + 1] if i + 1 < len(ordered) else None
            if nxt is not None:
                entry.bind("<Return>", lambda e, n=nxt: (n.focus_set(), "break"))
            else:
                entry.bind("<Return>", lambda e: (self.submit_business(), "break"))

    # ================================================================
    # Directorio (derecha)
    # ================================================================
    def _build_directory(self, parent):
        right = tk.Frame(parent, bg=SURFACE_ALT)
        right.pack(fill="both", expand=True)

        top = tk.Frame(right, bg=SURFACE_ALT)
        top.pack(fill="x", pady=(0, 8))
        tk.Label(top, text="Directorio de negocios", font=FONT_SUB, bg=SURFACE_ALT, fg=TEXT).pack(side="left")

        btn_bar = tk.Frame(top, bg=SURFACE_ALT)
        btn_bar.pack(side="right")

        import_btn = ModernButton(btn_bar, text="⬆ Importar", bg=SURFACE, fg=TEXT, hover=SURFACE_ALT,
                                   highlightbackground=BORDER, highlightthickness=1,
                                   command=self.import_data)
        import_btn.pack(side="left", padx=(0, 8))

        export_mb = tk.Menubutton(btn_bar, text="⬇ Exportar ▾", bg=SUCCESS, fg="white", relief="flat",
                                   bd=0, font=FONT_LABEL, padx=14, pady=8, cursor="hand2",
                                   activebackground="#0E6B4C", activeforeground="white")
        export_menu = tk.Menu(export_mb, tearoff=0)
        export_menu.add_command(label="Excel (.xlsx) — todo", command=self.export_excel)
        export_menu.add_command(label="CSV (.csv) — todo", command=self.export_csv)
        export_menu.add_command(label="PDF (.pdf) — todo", command=self.export_pdf)
        export_menu.add_separator()
        export_menu.add_command(label="Excel (.xlsx) — solo resultados filtrados",
                                 command=lambda: self.export_excel(filtered_only=True))
        export_menu.add_command(label="CSV (.csv) — solo resultados filtrados",
                                 command=lambda: self.export_csv(filtered_only=True))
        export_menu.add_command(label="PDF (.pdf) — solo resultados filtrados",
                                 command=lambda: self.export_pdf(filtered_only=True))
        export_mb.configure(menu=export_menu)
        export_mb.pack(side="left", padx=(8, 0))

        ModernButton(btn_bar, text="🖨️ Imprimir", bg=SURFACE, fg=TEXT, hover=SURFACE_ALT,
                     highlightbackground=BORDER, highlightthickness=1,
                     command=lambda: self.print_listing(filtered_only=False)).pack(side="left", padx=(8, 0))

        # ---- Barra de búsqueda y filtros ----
        toolbar = Card(right)
        toolbar.pack(fill="x", pady=(0, 10))
        trow = tk.Frame(toolbar, bg=SURFACE)
        trow.pack(fill="x", padx=12, pady=10)

        search_box = tk.Frame(trow, bg=SURFACE_ALT, highlightbackground=BORDER, highlightthickness=1)
        search_box.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(search_box, text="🔍", bg=SURFACE_ALT, fg=TEXT_FAINT).pack(side="left", padx=(8, 0))
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_box, textvariable=self.search_var, relief="flat", bd=0,
                                      bg=SURFACE_ALT, fg=TEXT, font=FONT_BASE)
        self.search_entry.pack(side="left", fill="x", expand=True, ipady=6, padx=6)
        self.search_entry.bind("<KeyRelease>", lambda e: self.refresh_notebook())

        self.filter_ciudad_var = tk.StringVar(value="Todas las ciudades")
        self.filter_ciudad_combo = ttk.Combobox(trow, textvariable=self.filter_ciudad_var, state="readonly",
                                                 width=16, values=["Todas las ciudades"])
        self.filter_ciudad_combo.pack(side="left", padx=4)
        self.filter_ciudad_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_notebook())

        self.filter_estado_var = tk.StringVar(value="Todos los estados")
        self.filter_estado_combo = ttk.Combobox(trow, textvariable=self.filter_estado_var, state="readonly",
                                                 width=15, values=["Todos los estados"] + ESTADOS)
        self.filter_estado_combo.pack(side="left", padx=4)
        self.filter_estado_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_notebook())

        self.only_fav_var = tk.BooleanVar(value=False)
        tk.Checkbutton(trow, text="⭐ Favoritos", variable=self.only_fav_var, bg=SURFACE, fg=TEXT_SOFT,
                        selectcolor=SURFACE, activebackground=SURFACE, font=FONT_SMALL,
                        command=self.refresh_notebook).pack(side="left", padx=6)

        self.sort_var = tk.StringVar(value="Nombre (A-Z)")
        self.sort_combo = ttk.Combobox(trow, textvariable=self.sort_var, state="readonly", width=20,
                                        values=["Nombre (A-Z)", "Código", "Más recientes primero",
                                                "Más antiguos primero"])
        self.sort_combo.pack(side="left", padx=(4, 0))
        self.sort_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_notebook())

        # ---- Notebook (Dashboard + Vista general + rubros) ----
        notebook_card = Card(right)
        notebook_card.pack(fill="both", expand=True)
        self.notebook = ttk.Notebook(notebook_card)
        self.notebook.pack(fill="both", expand=True, padx=4, pady=4)

        self.dashboard_frame = tk.Frame(self.notebook, bg=SURFACE)
        self.notebook.add(self.dashboard_frame, text="📊 Dashboard")

        # ---- Acciones sobre selección ----
        actions = tk.Frame(right, bg=SURFACE_ALT)
        actions.pack(fill="x", pady=(10, 0))
        ModernButton(actions, text="✏  Editar  (Ctrl+E)", bg=SURFACE, fg=TEXT, hover=SURFACE_ALT,
                     highlightbackground=BORDER, highlightthickness=1,
                     command=self.edit_selected).pack(side="left", padx=(0, 8))
        ModernButton(actions, text="⧉  Duplicar  (Ctrl+D)", bg=SURFACE, fg=TEXT, hover=SURFACE_ALT,
                     highlightbackground=BORDER, highlightthickness=1,
                     command=self.duplicate_selected).pack(side="left", padx=(0, 8))
        ModernButton(actions, text="🕒  Historial", bg=SURFACE, fg=TEXT, hover=SURFACE_ALT,
                     highlightbackground=BORDER, highlightthickness=1,
                     command=self.show_history_dialog).pack(side="left", padx=(0, 8))
        ModernButton(actions, text="🗑  Eliminar  (Supr)", bg=SURFACE, fg=DANGER, hover=DANGER_SOFT,
                     highlightbackground=BORDER, highlightthickness=1,
                     command=self.delete_selected).pack(side="left")

    # -----------------------------------------------------------------
    # Refresco de datos
    # -----------------------------------------------------------------
    def refresh_all(self, rebuild_tabs=False):
        self.categories = self.db.get_categories()
        self.refresh_rubro_combo()
        self.refresh_filter_options()
        self.refresh_notebook(rebuild_tabs=rebuild_tabs or self._categories_changed())
        self.scan_duplicate_notifications()

    def _categories_changed(self):
        return set(self.categories.keys()) != set(k for k in self.tab_widgets.keys())

    def refresh_filter_options(self):
        cities = self.db.distinct_cities()
        values = ["Todas las ciudades"] + cities
        self.filter_ciudad_combo["values"] = values
        if self.filter_ciudad_var.get() not in values:
            self.filter_ciudad_var.set("Todas las ciudades")

    def update_stats(self, total):
        self.stats_var.set(f"{len(self.categories)} rubros   ·   {total} negocios")

    def refresh_rubro_combo(self):
        names = list(self.categories.keys())
        values = [f"{n} — {self.categories[n]['prefix']}" for n in names] + ["+ Nueva categoría (rubro)"]
        current = self.rubro_var.get()
        self.rubro_combo["values"] = values
        if not current and values:
            self.rubro_combo.current(0)
        self.on_rubro_change()

    def selected_rubro_name(self):
        val = self.rubro_var.get()
        if val == "+ Nueva categoría (rubro)":
            return None
        return val.split(" — ")[0] if val else None

    def on_rubro_change(self, event=None):
        is_new = self.rubro_var.get() == "+ Nueva categoría (rubro)"
        if is_new:
            self.new_cat_frame.pack(fill="x", pady=(0, 10), after=self.rubro_combo)
        else:
            self.new_cat_frame.forget()
        self.update_code_preview()

    def update_code_preview(self):
        if self.rubro_var.get() == "+ Nueva categoría (rubro)":
            prefix = sanitize_prefix(self.new_cat_prefix.get())
            self.code_preview_var.set(f"{prefix}01" if prefix else "— · —")
        else:
            name = self.selected_rubro_name()
            cat = self.categories.get(name)
            if cat:
                self.code_preview_var.set(cat["prefix"] + pad_number(cat["counter"]))
            else:
                self.code_preview_var.set("—")

    # -----------------------------------------------------------------
    # Búsqueda / filtros / orden
    # -----------------------------------------------------------------
    def get_filtered_sorted(self, all_rows):
        query = self.search_var.get().strip().lower()
        ciudad_f = self.filter_ciudad_var.get()
        estado_f = self.filter_estado_var.get()
        only_fav = self.only_fav_var.get()
        sort_key = self.sort_var.get()

        def matches(b):
            if only_fav and not b["favorito"]:
                return False
            if ciudad_f != "Todas las ciudades" and (b["ciudad"] or "") != ciudad_f:
                return False
            if estado_f != "Todos los estados" and b["estado"] != estado_f:
                return False
            if query:
                query_digits = clean_phone_digits(query)
                haystack = " ".join([
                    b["negocio"] or "", b["code"] or "", b["category"] or "",
                    b["contacto"] or "", b["ciudad"] or "", b["tags"] or "",
                    b["telefono"] or "", b["correo"] or "", b["direccion"] or "",
                    b["pagina_web"] or "", b["ruc"] or "", b["cargo_contacto"] or "",
                ]).lower()
                # Si lo que se busca son puros dígitos (número de teléfono),
                # también se compara ignorando espacios/guiones/+ en el teléfono
                # guardado, para que "987654321" encuentre "+51 987 654 321".
                phone_match = bool(query_digits) and query_digits in clean_phone_digits(b["telefono"] or "")
                if query not in haystack and not phone_match:
                    return False
            return True

        rows = [b for b in all_rows if matches(b)]

        if sort_key == "Nombre (A-Z)":
            rows.sort(key=lambda b: (b["negocio"] or "").lower())
        elif sort_key == "Código":
            rows.sort(key=lambda b: b["code"])
        elif sort_key == "Más recientes primero":
            rows.sort(key=lambda b: parse_ddmmyyyy(b["fecha"]) or datetime.date.min, reverse=True)
        elif sort_key == "Más antiguos primero":
            rows.sort(key=lambda b: parse_ddmmyyyy(b["fecha"]) or datetime.date.min)
        return rows

    # -----------------------------------------------------------------
    # Notebook: Dashboard + Vista general + pestañas por rubro
    # -----------------------------------------------------------------
    def refresh_notebook(self, rebuild_tabs=False):
        all_rows, _ = self.db.get_businesses_grouped()
        filtered = self.get_filtered_sorted(all_rows)
        filtered_grouped = {}
        for row in filtered:
            filtered_grouped.setdefault(row["category"], []).append(row)

        if rebuild_tabs:
            self._rebuild_tabs(filtered_grouped, filtered)
        else:
            self._refresh_tab_contents(filtered, filtered_grouped)

        self.update_stats(len(all_rows))
        self.refresh_dashboard(all_rows)

    def _rebuild_tabs(self, grouped, all_filtered_rows):
        for tab_id in list(self.notebook.tabs()):
            frame_widget = self.nametowidget(tab_id)
            if frame_widget is self.dashboard_frame:
                continue
            self.notebook.forget(tab_id)
            frame_widget.destroy()

        self.tab_widgets = {}

        general_frame = tk.Frame(self.notebook, bg=SURFACE)
        self.notebook.add(general_frame, text="📋 Vista general")
        general_tree = self._build_tree(general_frame, GENERAL_COLUMNS, all_filtered_rows)
        self.tab_widgets["__general__"] = {"frame": general_frame, "tree": general_tree}

        for name, meta in self.categories.items():
            frame = tk.Frame(self.notebook, bg=SURFACE)
            self.notebook.add(frame, text=f"{name} ({meta['prefix']})")

            toolbar = tk.Frame(frame, bg=SURFACE)
            toolbar.pack(fill="x", padx=10, pady=(10, 0))
            tk.Label(toolbar, text=f"{name} · siglas {meta['prefix']}", font=FONT_LABEL,
                     bg=SURFACE, fg=TEXT).pack(side="left")
            ModernButton(toolbar, text="🗑 Eliminar rubro", bg=DANGER_SOFT, fg=DANGER, hover="#FEE4E2",
                         font=("Segoe UI", 8, "bold"), padx=8, pady=3,
                         command=lambda n=name: self.delete_category(n)).pack(side="right")

            rows = grouped.get(name, [])
            tree = self._build_tree(frame, COLUMNS, rows)
            self.tab_widgets[name] = {"frame": frame, "tree": tree}

    def _refresh_tab_contents(self, all_rows, grouped):
        general = self.tab_widgets.get("__general__")
        if general:
            self._fill_tree(general["tree"], GENERAL_COLUMNS, all_rows)
        for name, widgets in self.tab_widgets.items():
            if name == "__general__":
                continue
            rows = grouped.get(name, [])
            self._fill_tree(widgets["tree"], COLUMNS, rows)

    def _build_tree(self, parent, columns, rows):
        wrap = tk.Frame(parent, bg=SURFACE)
        wrap.pack(fill="both", expand=True, padx=10, pady=10)

        col_ids = [c[0] for c in columns]
        tree = ttk.Treeview(wrap, columns=col_ids, show="headings", selectmode="extended")
        for cid, heading, width in columns:
            tree.heading(cid, text=heading)
            tree.column(cid, width=width, anchor="w")

        for estado, color in ESTADO_COLORS.items():
            tree.tag_configure(f"estado_{estado}", foreground=color)

        tree._fav_col = "#" + str(col_ids.index("fav") + 1)

        vsb = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(wrap, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)

        tree.bind("<MouseWheel>", lambda e: tree.yview_scroll(int(-1 * (e.delta / 120)), "units"))
        tree.bind("<Button-4>", lambda e: tree.yview_scroll(-1, "units"))
        tree.bind("<Button-5>", lambda e: tree.yview_scroll(1, "units"))

        self._fill_tree(tree, columns, rows, wrap=wrap)
        tree.bind("<Double-1>", self._on_tree_double_click)
        tree.bind("<Button-1>", self._on_tree_click)
        tree.bind("<Button-3>", self._on_tree_right_click)
        parent.tree = tree
        return tree

    def _fill_tree(self, tree, columns, rows, wrap=None):
        col_ids = [c[0] for c in columns]
        selected = tree.selection()
        selected_codes = set(selected)

        tree.delete(*tree.get_children())
        for row in rows:
            values = []
            for cid in col_ids:
                if cid == "fav":
                    values.append("★" if row["favorito"] else "☆")
                else:
                    values.append(row[cid] if row[cid] is not None else "")
            tag = f"estado_{row['estado']}" if row["estado"] in ESTADO_COLORS else ""
            tree.insert("", "end", iid=row["code"], values=values, tags=(tag,) if tag else ())

        still_there = [c for c in selected_codes if tree.exists(c)]
        if still_there:
            tree.selection_set(still_there)

        wrap = wrap or tree.master
        empty_label = getattr(wrap, "_empty_label", None)
        if not rows:
            if empty_label is None:
                empty_label = tk.Label(wrap, text="No hay negocios que coincidan con la búsqueda o filtros.",
                                        bg=SURFACE, fg=TEXT_FAINT, font=FONT_SMALL)
                empty_label.grid(row=2, column=0, columnspan=2, pady=20)
                wrap._empty_label = empty_label
        elif empty_label is not None:
            empty_label.destroy()
            wrap._empty_label = None

    def get_active_tree(self):
        current = self.notebook.select()
        if not current:
            return None
        frame = self.nametowidget(current)
        return getattr(frame, "tree", None)

    def get_selected_codes(self):
        tree = self.get_active_tree()
        if not tree:
            return []
        return list(tree.selection())

    def get_selected_code(self):
        codes = self.get_selected_codes()
        return codes[0] if len(codes) == 1 else None

    # -----------------------------------------------------------------
    # Interacción con la tabla: clic en ★, doble clic, clic derecho
    # -----------------------------------------------------------------
    def _on_tree_click(self, event):
        tree = event.widget
        region = tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = tree.identify_column(event.x)
        row_id = tree.identify_row(event.y)
        if row_id and col == getattr(tree, "_fav_col", None):
            row = self.db.get_business(row_id)
            if row:
                self.db.set_favorite(row_id, not row["favorito"])
                self.db.log_history(row_id, row["negocio"], "Favorito",
                                     "Marcado como favorito." if not row["favorito"] else "Quitado de favoritos.")
                self.refresh_notebook()

    def _on_tree_double_click(self, event):
        tree = event.widget
        row_id = tree.identify_row(event.y)
        if row_id:
            tree.selection_set(row_id)
            self.open_edit_dialog(self.db.get_business(row_id))

    def _on_tree_right_click(self, event):
        tree = event.widget
        row_id = tree.identify_row(event.y)
        if row_id and row_id not in tree.selection():
            tree.selection_set(row_id)
        codes = list(tree.selection())
        if not codes:
            return
        self._context_menu_codes = codes
        menu = tk.Menu(self, tearoff=0)
        if len(codes) == 1:
            row = self.db.get_business(codes[0])
            menu.add_command(label=f"📋 Copiar teléfono ({row['telefono']})",
                              command=lambda: self.copy_to_clipboard(row["telefono"]))
            menu.add_command(label=f"📋 Copiar correo ({row['correo']})",
                              command=lambda: self.copy_to_clipboard(row["correo"]))
            menu.add_separator()
            menu.add_command(label="💬 Abrir WhatsApp", command=lambda: self.open_whatsapp(row))
            menu.add_command(label="📧 Redactar en Gmail", command=lambda: self.open_gmail(row))
            if (row["pagina_web"] or "").strip():
                menu.add_command(label="🌐 Abrir sitio web", command=lambda: self.open_website(row))
            menu.add_separator()
            self._add_move_submenu(menu, codes, current_category=row["category"])
            menu.add_command(label="✏ Editar", command=self.edit_selected)
            menu.add_command(label="⧉ Duplicar", command=self.duplicate_selected)
            menu.add_command(label="🕒 Ver historial", command=self.show_history_dialog)
            menu.add_separator()
            menu.add_command(label="🗑 Eliminar", command=self.delete_selected)
        else:
            self._add_move_submenu(menu, codes, current_category=None)
            menu.add_command(label=f"✏ Editar {len(codes)} registros en conjunto", command=self.edit_selected)
            menu.add_command(label="🗑 Eliminar seleccionados", command=self.delete_selected)
        menu.tk_popup(event.x_root, event.y_root)

    def _add_move_submenu(self, menu, codes, current_category):
        """Agrega al menú contextual la opción 'Mover a otro rubro', con un
        submenú listando los demás rubros disponibles."""
        other_names = [n for n in self.categories if n != current_category]
        if not other_names:
            return
        move_menu = tk.Menu(menu, tearoff=0)
        for name in other_names:
            move_menu.add_command(label=name, command=lambda n=name: self.move_selected_to_category(codes, n))
        menu.add_cascade(label="📁 Mover a otro rubro", menu=move_menu)
        menu.add_separator()

    def move_selected_to_category(self, codes, target_name):
        target = self.categories.get(target_name)
        if not target:
            return
        moved = 0
        for code in codes:
            row = self.db.get_business(code)
            if not row or row["category"] == target_name:
                continue
            self.db.move_category(code, target["id"])
            self.db.log_history(code, row["negocio"], "Cambio de rubro", f"{row['category']} → {target_name}")
            moved += 1
        if moved:
            self.refresh_all(rebuild_tabs=True)
            plural = "negocio" if moved == 1 else "negocios"
            self.show_msg(f"Se movió {moved} {plural} a \"{target_name}\". El código conserva su prefijo original.")

    # -----------------------------------------------------------------
    # Acciones rápidas de productividad
    # -----------------------------------------------------------------
    def copy_to_clipboard(self, text):
        self.clipboard_clear()
        self.clipboard_append(text or "")
        self.update()
        self.show_msg(f"Copiado al portapapeles: {text}")

    def open_whatsapp(self, row):
        digits = clean_phone_digits(row["telefono"])
        if not digits:
            return messagebox.showinfo("WhatsApp", "Este negocio no tiene un número de teléfono válido.")
        webbrowser.open(f"https://wa.me/{digits}")

    def open_gmail(self, row):
        correo = (row["correo"] or "").strip()
        if not correo:
            return messagebox.showinfo("Correo", "Este negocio no tiene un correo registrado.")
        webbrowser.open(f"https://mail.google.com/mail/?view=cm&fs=1&to={correo}")

    def open_website(self, row):
        url = ensure_url(row["pagina_web"])
        if not url:
            return messagebox.showinfo("Sitio web", "Este negocio no tiene una página web registrada.")
        webbrowser.open(url)

    # -----------------------------------------------------------------
    # Dashboard
    # -----------------------------------------------------------------
    def refresh_dashboard(self, businesses):
        frame = self.dashboard_frame
        if frame is None:
            return
        for w in frame.winfo_children():
            w.destroy()

        scroll = ScrollableFrame(frame, bg=SURFACE)
        scroll.pack(fill="both", expand=True)
        pad = scroll.inner
        pad.configure(padx=16, pady=16)

        total = len(businesses)
        favoritos = sum(1 for b in businesses if b["favorito"])
        today = datetime.date.today()
        vencidos = sum(1 for b in businesses
                        if (d := parse_ddmmyyyy(b["proximo_contacto"])) and d < today)
        hoy = sum(1 for b in businesses
                  if (d := parse_ddmmyyyy(b["proximo_contacto"])) and d == today)

        cards = tk.Frame(pad, bg=SURFACE)
        cards.pack(fill="x", pady=(0, 16))

        def stat_card(parent, title, value, color):
            c = Card(parent)
            c.pack(side="left", fill="both", expand=True, padx=(0, 10))
            tk.Label(c, text=str(value), font=("Segoe UI", 22, "bold"), bg=SURFACE, fg=color)\
                .pack(anchor="w", padx=14, pady=(12, 0))
            tk.Label(c, text=title, font=FONT_SMALL, bg=SURFACE, fg=TEXT_SOFT)\
                .pack(anchor="w", padx=14, pady=(0, 12))

        stat_card(cards, "Negocios totales", total, TEXT)
        stat_card(cards, "Rubros", len(self.categories), ACCENT)
        stat_card(cards, "Favoritos", favoritos, FAV_COLOR)
        stat_card(cards, "Seguimientos hoy", hoy, SUCCESS)
        stat_card(cards, "Seguimientos vencidos", vencidos, DANGER)

        cols = tk.Frame(pad, bg=SURFACE)
        cols.pack(fill="both", expand=True)

        left = Card(cols)
        left.pack(side="left", fill="both", expand=True, padx=(0, 8), anchor="n")
        tk.Label(left, text="Negocios por rubro", font=FONT_SUB, bg=SURFACE, fg=TEXT)\
            .pack(anchor="w", padx=14, pady=(12, 6))
        by_cat = {}
        for b in businesses:
            by_cat[b["category"]] = by_cat.get(b["category"], 0) + 1
        max_count = max(by_cat.values(), default=1)
        if by_cat:
            for name, count in sorted(by_cat.items(), key=lambda x: -x[1]):
                row = tk.Frame(left, bg=SURFACE)
                row.pack(fill="x", padx=14, pady=3)
                tk.Label(row, text=name, font=FONT_SMALL, bg=SURFACE, fg=TEXT, width=16, anchor="w")\
                    .pack(side="left")
                bar_bg = tk.Frame(row, bg=SURFACE_ALT, height=14)
                bar_bg.pack(side="left", fill="x", expand=True, padx=(4, 8))
                ratio = max(count / max_count, 0.04) if max_count else 0
                tk.Frame(bar_bg, bg=ACCENT, height=14).place(relx=0, rely=0, relwidth=ratio, relheight=1)
                tk.Label(row, text=str(count), font=FONT_SMALL, bg=SURFACE, fg=TEXT_SOFT, width=3)\
                    .pack(side="left")
        else:
            tk.Label(left, text="Todavía no hay negocios registrados.", font=FONT_SMALL,
                     bg=SURFACE, fg=TEXT_FAINT).pack(anchor="w", padx=14, pady=10)
        tk.Frame(left, bg=SURFACE, height=10).pack()

        by_estado = {}
        for b in businesses:
            by_estado[b["estado"]] = by_estado.get(b["estado"], 0) + 1
        if by_estado:
            tk.Label(left, text="Por estado", font=FONT_SUB, bg=SURFACE, fg=TEXT)\
                .pack(anchor="w", padx=14, pady=(4, 6))
            max_e = max(by_estado.values(), default=1)
            for name, count in by_estado.items():
                row = tk.Frame(left, bg=SURFACE)
                row.pack(fill="x", padx=14, pady=3)
                tk.Label(row, text=name, font=FONT_SMALL, bg=SURFACE,
                         fg=ESTADO_COLORS.get(name, TEXT), width=16, anchor="w").pack(side="left")
                bar_bg = tk.Frame(row, bg=SURFACE_ALT, height=14)
                bar_bg.pack(side="left", fill="x", expand=True, padx=(4, 8))
                ratio = max(count / max_e, 0.04) if max_e else 0
                tk.Frame(bar_bg, bg=ESTADO_COLORS.get(name, ACCENT), height=14)\
                    .place(relx=0, rely=0, relwidth=ratio, relheight=1)
                tk.Label(row, text=str(count), font=FONT_SMALL, bg=SURFACE, fg=TEXT_SOFT, width=3)\
                    .pack(side="left")
        tk.Frame(left, bg=SURFACE, height=10).pack()

        right = tk.Frame(cols, bg=SURFACE)
        right.pack(side="left", fill="both", expand=True, padx=(8, 0))

        upcoming = Card(right)
        upcoming.pack(fill="both", expand=True, pady=(0, 8))
        tk.Label(upcoming, text="Próximos seguimientos (7 días)", font=FONT_SUB, bg=SURFACE, fg=TEXT)\
            .pack(anchor="w", padx=14, pady=(12, 6))
        upcoming_list = []
        for b in businesses:
            d = parse_ddmmyyyy(b["proximo_contacto"])
            if d and today <= d <= today + datetime.timedelta(days=7):
                upcoming_list.append((d, b))
        upcoming_list.sort(key=lambda x: x[0])
        if upcoming_list:
            for d, b in upcoming_list[:10]:
                tk.Label(upcoming, text=f"{d.strftime('%d/%m')} · {b['negocio']} ({b['code']})",
                         font=FONT_SMALL, bg=SURFACE, fg=TEXT, anchor="w").pack(anchor="w", padx=14, pady=1)
        else:
            tk.Label(upcoming, text="No hay seguimientos programados en los próximos 7 días.",
                     font=FONT_SMALL, bg=SURFACE, fg=TEXT_FAINT).pack(anchor="w", padx=14, pady=6)
        tk.Frame(upcoming, bg=SURFACE, height=10).pack()

        activity = Card(right)
        activity.pack(fill="both", expand=True)
        tk.Label(activity, text="Actividad reciente", font=FONT_SUB, bg=SURFACE, fg=TEXT)\
            .pack(anchor="w", padx=14, pady=(12, 6))
        history = self.db.get_history(limit=10)
        if history:
            for h in history:
                tk.Label(activity, text=f"{h['timestamp']} · {h['action']} · {h['negocio']} ({h['code']})",
                         font=("Segoe UI", 8), bg=SURFACE, fg=TEXT_SOFT, anchor="w", justify="left")\
                    .pack(anchor="w", padx=14, pady=1, fill="x")
        else:
            tk.Label(activity, text="Sin actividad registrada todavía.", font=FONT_SMALL,
                     bg=SURFACE, fg=TEXT_FAINT).pack(anchor="w", padx=14, pady=6)
        tk.Frame(activity, bg=SURFACE, height=10).pack()

    # -----------------------------------------------------------------
    # Registrar negocio
    # -----------------------------------------------------------------
    def _collect_form_data(self):
        data = {k: e.get().strip() for k, e in self.fields.items()}
        data["notas"] = self.notas_text.get("1.0", "end").strip()
        data["estado"] = self.estado_var.get()
        data["ultimo_contacto"] = self.ultimo_contacto_entry.get().strip()
        data["proximo_contacto"] = self.proximo_contacto_entry.get().strip()
        data["favorito"] = self.favorito_var.get()
        return data

    def submit_business(self):
        self.msg_var.set("")
        rubro_val = self.rubro_var.get()

        if rubro_val == "+ Nueva categoría (rubro)":
            name = self.new_cat_name.get().strip()
            prefix = sanitize_prefix(self.new_cat_prefix.get())
            if not name or not prefix:
                return self.show_msg("Indica el nombre del rubro y sus siglas antes de continuar.", error=True)
            existing = next((k for k in self.categories if k.lower() == name.lower()), None)
            if existing:
                cat_name = existing
            else:
                self.db.create_category(name, prefix)
                self.categories = self.db.get_categories()
                cat_name = name
        else:
            cat_name = self.selected_rubro_name()
            if not cat_name:
                return self.show_msg("Selecciona un rubro.", error=True)

        data = self._collect_form_data()

        if not data["negocio"] or not data["telefono"] or not data["correo"] or not data["direccion"]:
            return self.show_msg("Nombre del negocio, número de contacto, correo y dirección son obligatorios.",
                                  error=True)

        cat = self.categories[cat_name]
        code = cat["prefix"] + pad_number(cat["counter"])
        self.db.insert_business(code, cat["id"], data)
        self.db.bump_counter(cat["id"], cat["counter"] + 1)

        for e in self.fields.values():
            e.delete(0, "end")
        self.notas_text.delete("1.0", "end")
        self.estado_var.set("Activo")
        self.ultimo_contacto_entry.delete(0, "end")
        self.proximo_contacto_entry.delete(0, "end")
        self.favorito_var.set(False)

        new_category = cat_name not in self.tab_widgets
        self.refresh_all(rebuild_tabs=new_category)
        self.show_msg(f"Negocio registrado con el código {code}.")
        self.fields["negocio"].focus_set()

    def show_msg(self, text, error=False):
        self.msg_var.set(text)
        self.msg_label.configure(fg=DANGER if error else SUCCESS)

    # -----------------------------------------------------------------
    # Editar (individual o en conjunto) / Duplicar / Eliminar
    # -----------------------------------------------------------------
    def edit_selected(self):
        codes = self.get_selected_codes()
        if not codes:
            return messagebox.showinfo("Editar", "Selecciona primero uno o varios negocios en la tabla.")
        if len(codes) == 1:
            row = self.db.get_business(codes[0])
            if row:
                self.open_edit_dialog(row)
        else:
            self.open_bulk_edit_dialog(codes)

    def open_edit_dialog(self, row, mode="editar"):
        dialog = tk.Toplevel(self)
        dialog.title(f"{'Duplicar' if mode == 'duplicar' else 'Editar'} negocio — {row['code']}")
        dialog.configure(bg=SURFACE)
        dialog.geometry("440x760")
        dialog.minsize(380, 480)
        dialog.transient(self)
        dialog.grab_set()
        dialog.bind("<Escape>", lambda e: dialog.destroy())

        badge_text = f"{row['code']} · {row['category']}"
        if mode == "duplicar":
            badge_text += "  (se generará un código nuevo)"
        tk.Label(dialog, text=badge_text, font=("Consolas", 10, "bold"),
                 bg=ACCENT_SOFT, fg=ACCENT, wraplength=400, justify="left")\
            .pack(fill="x", padx=16, pady=(16, 10), ipady=6)

        scroll = ScrollableFrame(dialog, bg=SURFACE)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)
        inner = scroll.inner
        inner.configure(padx=16)

        entries = {}
        ordered_entries = []
        field_defs = [
            ("negocio", "Nombre del negocio *"),
            ("contacto", "Nombre de contacto"),
            ("telefono", "Número de contacto *"),
            ("correo", "Correo electrónico *"),
            ("direccion", "Dirección *"),
            ("ciudad", "Ciudad"),
            ("pagina_web", "Página web"),
            ("cargo_contacto", "Cargo de contacto"),
            ("ruc", "RUC / NIT / CIF"),
            ("tags", "Etiquetas (separadas por coma)"),
        ]
        for key, label in field_defs:
            tk.Label(inner, text=label, font=FONT_LABEL, bg=SURFACE, fg=TEXT_SOFT)\
                .pack(anchor="w", pady=(10, 2))
            entry = styled_entry(inner)
            entry.insert(0, row[key] or "")
            entry.pack(fill="x", pady=2, ipady=4)
            entries[key] = entry
            ordered_entries.append(entry)

        tk.Label(inner, text="Estado del cliente", font=FONT_LABEL, bg=SURFACE, fg=TEXT_SOFT)\
            .pack(anchor="w", pady=(10, 2))
        estado_var = tk.StringVar(value=row["estado"])
        ttk.Combobox(inner, textvariable=estado_var, values=ESTADOS, state="readonly", font=FONT_BASE)\
            .pack(fill="x", pady=2, ipady=3)

        ultimo_entry = build_date_field(inner, "Último contacto",
                                         "" if mode == "duplicar" else (row["ultimo_contacto"] or ""))
        proximo_entry = build_date_field(inner, "Próxima fecha de seguimiento",
                                          "" if mode == "duplicar" else (row["proximo_contacto"] or ""))

        favorito_var = tk.BooleanVar(value=bool(row["favorito"]) if mode != "duplicar" else False)
        tk.Checkbutton(inner, text="⭐ Marcar como favorito", variable=favorito_var, bg=SURFACE,
                        fg=TEXT_SOFT, selectcolor=SURFACE, activebackground=SURFACE, font=FONT_SMALL)\
            .pack(anchor="w", pady=(10, 0))

        tk.Label(inner, text="Notas", font=FONT_LABEL, bg=SURFACE, fg=TEXT_SOFT).pack(anchor="w", pady=(10, 2))
        notas_text = tk.Text(inner, height=4, relief="flat", bd=0, font=FONT_SMALL, bg=SURFACE_ALT,
                              fg=TEXT, highlightthickness=1, highlightbackground=BORDER,
                              highlightcolor=ACCENT, wrap="word")
        notas_text.insert("1.0", "" if mode == "duplicar" else (row["notas"] or ""))
        notas_text.pack(fill="x", pady=(2, 16))

        def save(event=None):
            data = {k: e.get().strip() for k, e in entries.items()}
            data["estado"] = estado_var.get()
            data["notas"] = notas_text.get("1.0", "end").strip()
            data["ultimo_contacto"] = ultimo_entry.get().strip()
            data["proximo_contacto"] = proximo_entry.get().strip()
            data["favorito"] = favorito_var.get()
            if not data["negocio"] or not data["telefono"] or not data["correo"] or not data["direccion"]:
                messagebox.showerror("Datos incompletos",
                                      "Nombre del negocio, número de contacto, correo y dirección son obligatorios.",
                                      parent=dialog)
                return
            if mode == "duplicar":
                cat_name = row["category"]
                cat = self.categories[cat_name]
                new_code = cat["prefix"] + pad_number(cat["counter"])
                self.db.insert_business(new_code, cat["id"], data)
                self.db.bump_counter(cat["id"], cat["counter"] + 1)
                self.db.log_history(new_code, data["negocio"], "Duplicado", f"Copiado a partir de {row['code']}.")
                dialog.destroy()
                self.refresh_all()
                self.show_msg(f"Se creó la copia con el código {new_code}.")
            else:
                self.db.update_business(row["code"], data)
                dialog.destroy()
                self.refresh_all()
                self.show_msg(f"Datos de {data['negocio']} ({row['code']}) actualizados.")

        for i, entry in enumerate(ordered_entries):
            if i + 1 < len(ordered_entries):
                nxt = ordered_entries[i + 1]
                entry.bind("<Return>", lambda e, n=nxt: (n.focus_set(), "break"))
            else:
                entry.bind("<Return>", lambda e: (save(), "break"))
        dialog.bind("<Control-Return>", save)

        btns = tk.Frame(dialog, bg=SURFACE)
        btns.pack(fill="x", padx=16, pady=16)
        ModernButton(btns, text="Cancelar", bg=SURFACE, fg=TEXT_SOFT, hover=SURFACE_ALT,
                     highlightbackground=BORDER, highlightthickness=1,
                     command=dialog.destroy).pack(side="left", expand=True, fill="x", padx=(0, 8))
        save_label = "Guardar copia  (Ctrl+Enter)" if mode == "duplicar" else "Guardar cambios  (Ctrl+Enter)"
        ModernButton(btns, text=save_label, bg=ACCENT, fg="white", command=save)\
            .pack(side="left", expand=True, fill="x")

        entries["negocio"].focus_set()

    def duplicate_selected(self):
        codes = self.get_selected_codes()
        if len(codes) != 1:
            return messagebox.showinfo("Duplicar", "Selecciona exactamente un negocio para duplicarlo.")
        row = self.db.get_business(codes[0])
        if row:
            self.open_edit_dialog(row, mode="duplicar")

    # -----------------------------------------------------------------
    # Edición masiva (varios seleccionados)
    # -----------------------------------------------------------------
    def open_bulk_edit_dialog(self, codes):
        dialog = tk.Toplevel(self)
        dialog.title(f"Editar {len(codes)} negocios en conjunto")
        dialog.configure(bg=SURFACE)
        dialog.geometry("420x480")
        dialog.transient(self)
        dialog.grab_set()
        dialog.bind("<Escape>", lambda e: dialog.destroy())

        tk.Label(dialog, text=f"Se aplicará solo a los campos que completes,\na los {len(codes)} registros seleccionados.",
                 font=("Segoe UI", 9), bg=WARN_SOFT, fg=WARN_TEXT, justify="left")\
            .pack(fill="x", padx=16, pady=(16, 10), ipady=8)

        body = tk.Frame(dialog, bg=SURFACE)
        body.pack(fill="both", expand=True, padx=16)

        tk.Label(body, text="Estado (dejar en blanco = sin cambios)", font=FONT_LABEL, bg=SURFACE,
                 fg=TEXT_SOFT).pack(anchor="w", pady=(6, 2))
        estado_var = tk.StringVar(value="")
        ttk.Combobox(body, textvariable=estado_var, values=[""] + ESTADOS, state="readonly")\
            .pack(fill="x", pady=2, ipady=3)

        tk.Label(body, text="Ciudad (dejar en blanco = sin cambios)", font=FONT_LABEL, bg=SURFACE,
                 fg=TEXT_SOFT).pack(anchor="w", pady=(10, 2))
        ciudad_entry = styled_entry(body)
        ciudad_entry.pack(fill="x", pady=2, ipady=4)

        tk.Label(body, text="Favorito", font=FONT_LABEL, bg=SURFACE, fg=TEXT_SOFT)\
            .pack(anchor="w", pady=(10, 2))
        fav_var = tk.StringVar(value="Sin cambios")
        ttk.Combobox(body, textvariable=fav_var, state="readonly",
                     values=["Sin cambios", "Marcar como favorito", "Quitar de favoritos"])\
            .pack(fill="x", pady=2, ipady=3)

        tk.Label(body, text="Agregar etiqueta (se suma a las que ya tenga cada uno)", font=FONT_LABEL,
                 bg=SURFACE, fg=TEXT_SOFT).pack(anchor="w", pady=(10, 2))
        tag_entry = styled_entry(body)
        tag_entry.pack(fill="x", pady=2, ipady=4)

        proximo_entry = build_date_field(body, "Próxima fecha de seguimiento (blanco = sin cambios)")

        def save():
            changes = {}
            if estado_var.get():
                changes["estado"] = estado_var.get()
            if ciudad_entry.get().strip():
                changes["ciudad"] = ciudad_entry.get().strip()
            if proximo_entry.get().strip():
                changes["proximo_contacto"] = proximo_entry.get().strip()
            if fav_var.get() == "Marcar como favorito":
                changes["favorito"] = 1
            elif fav_var.get() == "Quitar de favoritos":
                changes["favorito"] = 0

            new_tag = tag_entry.get().strip()

            if not changes and not new_tag:
                messagebox.showinfo("Sin cambios", "No completaste ningún campo para aplicar.", parent=dialog)
                return

            if changes:
                self.db.bulk_update(codes, changes)
            if new_tag:
                for code in codes:
                    row = self.db.get_business(code)
                    existing = [t.strip() for t in (row["tags"] or "").split(",") if t.strip()]
                    if new_tag not in existing:
                        existing.append(new_tag)
                    self.db.set_tags(code, ", ".join(existing))

            summary = "; ".join(f"{FIELD_LABELS.get(k, k)} → {v}" for k, v in changes.items())
            if new_tag:
                summary += (("; " if summary else "") + f"etiqueta añadida: {new_tag}")
            for code in codes:
                row = self.db.get_business(code)
                self.db.log_history(code, row["negocio"] if row else "", "Edición masiva", summary)

            dialog.destroy()
            self.refresh_all()
            self.show_msg(f"Se actualizaron {len(codes)} registros.")

        btns = tk.Frame(dialog, bg=SURFACE)
        btns.pack(fill="x", padx=16, pady=16)
        ModernButton(btns, text="Cancelar", bg=SURFACE, fg=TEXT_SOFT, hover=SURFACE_ALT,
                     highlightbackground=BORDER, highlightthickness=1,
                     command=dialog.destroy).pack(side="left", expand=True, fill="x", padx=(0, 8))
        ModernButton(btns, text="Aplicar cambios", bg=ACCENT, fg="white", command=save)\
            .pack(side="left", expand=True, fill="x")

    def delete_selected(self):
        codes = self.get_selected_codes()
        if not codes:
            return messagebox.showinfo("Eliminar", "Selecciona primero uno o varios negocios en la tabla.")
        if len(codes) == 1:
            row = self.db.get_business(codes[0])
            question = f"¿Eliminar el registro de \"{row['negocio']}\" ({codes[0]})?"
        else:
            question = f"¿Eliminar los {len(codes)} registros seleccionados? Esta acción no se puede deshacer."
        if not messagebox.askyesno("Eliminar negocio(s)", question):
            return
        for code in codes:
            self.db.delete_business(code)
        self.refresh_all()

    # -----------------------------------------------------------------
    # Historial de modificaciones
    # -----------------------------------------------------------------
    def show_history_dialog(self):
        codes = self.get_selected_codes()
        code = codes[0] if len(codes) == 1 else None

        dialog = tk.Toplevel(self)
        dialog.title(f"Historial — {code}" if code else "Historial de actividad (global)")
        dialog.configure(bg=SURFACE)
        dialog.geometry("640x460")
        dialog.transient(self)
        dialog.grab_set()
        dialog.bind("<Escape>", lambda e: dialog.destroy())

        tk.Label(dialog, text="Historial de modificaciones" + (f" de {code}" if code else " (todos los negocios)"),
                 font=FONT_SUB, bg=SURFACE, fg=TEXT).pack(anchor="w", padx=16, pady=(16, 8))

        wrap = tk.Frame(dialog, bg=SURFACE)
        wrap.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        cols = ("timestamp", "code", "negocio", "action", "detail")
        tree = ttk.Treeview(wrap, columns=cols, show="headings")
        headers = {"timestamp": "Fecha/hora", "code": "Código", "negocio": "Negocio",
                   "action": "Acción", "detail": "Detalle"}
        widths = {"timestamp": 110, "code": 70, "negocio": 140, "action": 90, "detail": 260}
        for c in cols:
            tree.heading(c, text=headers[c])
            tree.column(c, width=widths[c], anchor="w")
        vsb = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="left", fill="y")

        rows = self.db.get_history(code=code, limit=300)
        for h in rows:
            tree.insert("", "end", values=(h["timestamp"], h["code"], h["negocio"], h["action"], h["detail"]))
        if not rows:
            tk.Label(dialog, text="Todavía no hay actividad registrada.", bg=SURFACE, fg=TEXT_FAINT,
                     font=FONT_SMALL).pack(pady=10)

    # -----------------------------------------------------------------
    # Notificaciones (duplicados detectados, avisos de importación, etc.)
    # -----------------------------------------------------------------
    def scan_duplicate_notifications(self):
        """Recalcula los avisos de negocios con el mismo nombre que ya existen
        en la base de datos. No bloquea nada — solo lo deja anotado para
        revisarlo cuando quieras (los registros se mantienen tal cual)."""
        self.notifications = [n for n in self.notifications if n.get("kind") != "duplicado_existente"]
        now = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        for group in self.db.duplicate_name_groups():
            codes = ", ".join(g["code"] for g in group)
            nombre = group[0]["negocio"]
            self.notifications.append({
                "kind": "duplicado_existente",
                "text": f"Posible duplicado: \"{nombre}\" aparece en {len(group)} registros ({codes}). "
                        f"Revisa si son el mismo negocio.",
                "timestamp": now,
            })
        self._update_notif_badge()

    def add_import_duplicate_notification(self, nombre, code_existente):
        self.notifications.append({
            "kind": "duplicado_importado",
            "text": f"Al importar, \"{nombre}\" ya existía como {code_existente} — no se creó un registro "
                    f"nuevo para evitar el duplicado.",
            "timestamp": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
        })
        self._update_notif_badge()

    def _update_notif_badge(self):
        if not hasattr(self, "notif_btn") or not self.notif_btn.winfo_exists():
            return
        count = len(self.notifications)
        self.notif_btn.configure(text=f"🔔 Notificaciones ({count})" if count else "🔔 Notificaciones")

    def show_notifications_dialog(self):
        dialog = tk.Toplevel(self)
        dialog.title("Notificaciones")
        dialog.configure(bg=SURFACE)
        dialog.geometry("560x440")
        dialog.minsize(360, 300)
        dialog.transient(self)
        dialog.bind("<Escape>", lambda e: dialog.destroy())

        tk.Label(dialog, text="Notificaciones", font=FONT_SUB, bg=SURFACE, fg=TEXT)\
            .pack(anchor="w", padx=16, pady=(16, 4))
        tk.Label(dialog, text="Avisos para revisar — no bloquean tu trabajo, los registros se mantienen.",
                 font=FONT_SMALL, bg=SURFACE, fg=TEXT_FAINT).pack(anchor="w", padx=16, pady=(0, 8))

        scroll = ScrollableFrame(dialog, bg=SURFACE)
        scroll.pack(fill="both", expand=True, padx=16)
        inner = scroll.inner
        inner.configure(pady=4)

        if not self.notifications:
            tk.Label(inner, text="No hay notificaciones pendientes.", font=FONT_SMALL,
                     bg=SURFACE, fg=TEXT_FAINT).pack(anchor="w", pady=10)
        else:
            for n in reversed(self.notifications):
                card = tk.Frame(inner, bg=WARN_SOFT, highlightbackground="#F2C94C", highlightthickness=1)
                card.pack(fill="x", pady=4)
                tk.Label(card, text=n["text"], font=FONT_SMALL, bg=WARN_SOFT, fg=WARN_TEXT,
                         wraplength=480, justify="left").pack(anchor="w", padx=10, pady=(8, 2))
                tk.Label(card, text=n["timestamp"], font=("Segoe UI", 7), bg=WARN_SOFT, fg=TEXT_FAINT)\
                    .pack(anchor="w", padx=10, pady=(0, 8))

        btns = tk.Frame(dialog, bg=SURFACE)
        btns.pack(fill="x", padx=16, pady=12)

        def clear_import_notifications():
            self.notifications = [n for n in self.notifications if n.get("kind") != "duplicado_importado"]
            self._update_notif_badge()
            dialog.destroy()
            self.show_notifications_dialog()

        ModernButton(btns, text="Descartar avisos de importación", bg=SURFACE, fg=TEXT, hover=SURFACE_ALT,
                     highlightbackground=BORDER, highlightthickness=1,
                     command=clear_import_notifications).pack(side="left", expand=True, fill="x", padx=(0, 8))
        ModernButton(btns, text="Cerrar", bg=ACCENT, fg="white", command=dialog.destroy)\
            .pack(side="left", expand=True, fill="x")

    # -----------------------------------------------------------------
    # Eliminar rubro
    # -----------------------------------------------------------------
    def delete_category(self, name):
        meta = self.categories.get(name)
        if not meta:
            return
        _all_rows, grouped = self.db.get_businesses_grouped()
        count = len(grouped.get(name, []))
        if count > 0:
            question = (f"El rubro \"{name}\" tiene {count} negocio(s) registrado(s). "
                        f"Si lo eliminas, también se eliminarán esos registros.\n\n¿Deseas continuar?")
        else:
            question = f"¿Eliminar el rubro \"{name}\"?"
        if not messagebox.askyesno("Eliminar rubro", question):
            return
        self.db.delete_category(meta["id"])
        self.rubro_var.set("")
        self.refresh_all(rebuild_tabs=True)

    # -----------------------------------------------------------------
    # Importar desde Excel / CSV
    # -----------------------------------------------------------------
    def import_data(self):
        path = filedialog.askopenfilename(
            title="Importar negocios desde Excel o CSV",
            filetypes=[("Excel o CSV", "*.xlsx *.xls *.csv"), ("Excel", "*.xlsx *.xls"), ("CSV", "*.csv")]
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext in (".xlsx", ".xls"):
                if not OPENPYXL_OK:
                    messagebox.showerror("Falta un paquete",
                                          "Para importar Excel necesitas instalar openpyxl.\n\npip install openpyxl")
                    return
                stats = self._import_xlsx(path)
            elif ext == ".csv":
                stats = self._import_csv(path)
            else:
                messagebox.showerror("Formato no soportado", "Solo se pueden importar archivos .xlsx o .csv")
                return
        except Exception as exc:
            messagebox.showerror("Error al importar", f"No se pudo importar el archivo:\n{exc}")
            return

        self.refresh_all(rebuild_tabs=True)
        msg = (f"Importación completada:\n"
               f"  • {stats['nuevos']} negocio(s) nuevo(s)\n"
               f"  • {stats['actualizados']} negocio(s) actualizado(s) (el código ya existía)\n"
               f"  • {stats['rubros_creados']} rubro(s) nuevo(s) creado(s)")
        if stats.get("duplicados"):
            msg += (f"\n  • {stats['duplicados']} fila(s) omitida(s) por nombre duplicado "
                     f"(ya revisables en 🔔 Notificaciones)")
        if stats["errores"]:
            msg += f"\n\nFilas omitidas por falta de nombre: {stats['errores']}"
        messagebox.showinfo("Importación completada", msg)

    def _get_or_create_category(self, name, stats):
        name = (name or "").strip() or "Sin rubro"
        existing = next((k for k in self.categories if k.lower() == name.lower()), None)
        if existing:
            return self.categories[existing]
        prefix = sanitize_prefix(name)[:4] or "GEN"
        used_prefixes = {m["prefix"] for m in self.categories.values()}
        base_prefix, n = prefix, 2
        while prefix in used_prefixes:
            prefix = (base_prefix[:3] + str(n))[:4]
            n += 1
        self.db.create_category(name, prefix)
        self.categories = self.db.get_categories()
        stats["rubros_creados"] += 1
        return self.categories[name]

    def _import_row(self, rubro_name, row_data, stats):
        negocio = (row_data.get("negocio") or "").strip()
        # Solo el nombre es realmente indispensable: los directorios tipo
        # Google Maps casi nunca traen correo, y no queremos descartar filas
        # válidas solo por eso. telefono/correo/direccion quedan en blanco
        # si el archivo no los trae, sin romper la estructura de columnas.
        if not negocio:
            stats["errores"] += 1
            return

        cat = self._get_or_create_category(rubro_name, stats)
        data = {k: (row_data.get(k) or "") for k in FULL_FIELDS}
        if data.get("estado") not in ESTADOS:
            data["estado"] = "Activo"

        code = (row_data.get("code") or "").strip().upper()
        if code and self.db.code_exists(code):
            self.db.update_business(code, data, log=False)
            self.db.log_history(code, negocio, "Importado", "Actualizado desde archivo importado.")
            stats["actualizados"] += 1
            return

        # Sin código que empate: evitar crear un negocio duplicado si ya
        # existe uno con el mismo nombre (sin importar mayúsculas/espacios).
        existing = self.db.find_by_name(negocio)
        if existing is not None:
            stats["duplicados"] += 1
            self.add_import_duplicate_notification(negocio, existing["code"])
            return

        if not code or self.db.code_exists(code):
            code = cat["prefix"] + pad_number(cat["counter"])
        while self.db.code_exists(code):
            cat["counter"] += 1
            code = cat["prefix"] + pad_number(cat["counter"])
        self.db.insert_business(code, cat["id"], data, log=False)
        cat["counter"] += 1
        self.db.bump_counter(cat["id"], cat["counter"])
        self.db.log_history(code, negocio, "Importado", "Creado desde archivo importado.")
        stats["nuevos"] += 1

    def _import_xlsx(self, path):
        stats = {"nuevos": 0, "actualizados": 0, "rubros_creados": 0, "errores": 0, "duplicados": 0}
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet in wb.worksheets:
            if sheet.title.strip().lower() == "resumen":
                continue
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue
            header_map, has_rubro_col = {}, False
            for idx, h in enumerate(header_row):
                if h is None:
                    continue
                key = normalize_header(str(h))
                if key:
                    header_map[idx] = key
                    if key == "rubro":
                        has_rubro_col = True
            if not header_map:
                continue
            for values in rows_iter:
                if values is None or all(v is None for v in values):
                    continue
                row_data = {}
                for idx, key in header_map.items():
                    if idx < len(values):
                        val = values[idx]
                        row_data[key] = "" if val is None else str(val).strip()
                rubro_name = row_data.get("rubro") if has_rubro_col else sheet.title
                self._import_row(rubro_name, row_data, stats)
        return stats

    def _ask_default_rubro(self):
        from tkinter import simpledialog
        return simpledialog.askstring(
            "Rubro para este archivo",
            "El archivo no tiene una columna de Rubro/Categoría.\n"
            "Escribe el nombre del rubro al que pertenecen todos estos negocios:",
            parent=self
        )

    def _import_csv(self, path):
        stats = {"nuevos": 0, "actualizados": 0, "rubros_creados": 0, "errores": 0, "duplicados": 0}
        with open(path, newline="", encoding="utf-8-sig") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;")
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(f, dialect=dialect)
            if not reader.fieldnames:
                return stats
            header_map, has_rubro_col = {}, False
            for h in reader.fieldnames:
                key = normalize_header(h)
                if key:
                    header_map[h] = key
                    if key == "rubro":
                        has_rubro_col = True
            default_rubro = None
            if not has_rubro_col:
                default_rubro = self._ask_default_rubro()
                if not default_rubro:
                    return stats
            for raw_row in reader:
                row_data = {}
                for h, key in header_map.items():
                    row_data[key] = (raw_row.get(h) or "").strip()
                rubro_name = row_data.get("rubro") if has_rubro_col else default_rubro
                self._import_row(rubro_name, row_data, stats)
        return stats

    # -----------------------------------------------------------------
    # Exportar a Excel
    # -----------------------------------------------------------------
    EXPORT_HEADERS = ["Código", "Nombre del Negocio", "Nombre de Contacto", "Número de Contacto",
                       "Correo Electrónico", "Dirección", "Ciudad", "Estado", "Fecha de Registro",
                       "Página Web", "Cargo de Contacto", "RUC/NIT/CIF", "Favorito", "Etiquetas",
                       "Último Contacto", "Próximo Seguimiento", "Notas"]

    @staticmethod
    def _export_row(b):
        return [b["code"], b["negocio"], b["contacto"], b["telefono"], b["correo"], b["direccion"],
                b["ciudad"], b["estado"], b["fecha"], b["pagina_web"], b["cargo_contacto"], b["ruc"],
                "Sí" if b["favorito"] else "No", b["tags"], b["ultimo_contacto"], b["proximo_contacto"],
                b["notas"]]

    def _grouped_for_export(self, filtered_only):
        """Devuelve (filas, agrupado_por_rubro) — todo el directorio, o solo
        lo que coincide con la búsqueda/filtros activos en pantalla."""
        all_rows, grouped = self.db.get_businesses_grouped()
        if not filtered_only:
            return all_rows, grouped
        filtered = self.get_filtered_sorted(all_rows)
        filtered_grouped = {}
        for row in filtered:
            filtered_grouped.setdefault(row["category"], []).append(row)
        return filtered, filtered_grouped

    def export_excel(self, filtered_only=False):
        if not OPENPYXL_OK:
            messagebox.showerror(
                "Falta un paquete",
                "Para exportar a Excel necesitas instalar openpyxl.\n\nAbre una terminal y ejecuta:\n\npip install openpyxl"
            )
            return

        default_name = "Directorio_Negocios_filtrado.xlsx" if filtered_only else "Directorio_Negocios.xlsx"
        path = filedialog.asksaveasfilename(
            title="Guardar directorio de negocios", defaultextension=".xlsx",
            initialfile=default_name, filetypes=[("Libro de Excel", "*.xlsx")]
        )
        if not path:
            return

        businesses, grouped = self._grouped_for_export(filtered_only)
        wb = openpyxl.Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

        ws_resumen.append(["Rubro", "Siglas", "Negocios registrados", "Próximo código"])
        for c in ws_resumen[1]:
            c.font = header_font
            c.fill = header_fill
        for name, meta in self.categories.items():
            count = len(grouped.get(name, []))
            ws_resumen.append([name, meta["prefix"], count, meta["prefix"] + pad_number(meta["counter"])])
        for i, w in enumerate([22, 12, 20, 16], start=1):
            ws_resumen.column_dimensions[get_column_letter(i)].width = w

        used_names = {"Resumen"}
        for name in self.categories:
            safe = "".join(ch for ch in name if ch not in '[]*?/\\:')[:31] or "Rubro"
            final, i = safe, 2
            while final in used_names:
                final = f"{safe[:28]}_{i}"
                i += 1
            used_names.add(final)

            ws = wb.create_sheet(final)
            ws.append(self.EXPORT_HEADERS)
            for c in ws[1]:
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center", wrap_text=True)
            rows = grouped.get(name, [])
            for b in rows:
                ws.append(self._export_row(b))
            widths = [10, 24, 18, 16, 26, 26, 12, 10, 14, 22, 16, 14, 9, 18, 13, 15, 26]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(self.EXPORT_HEADERS))}{max(1, len(rows) + 1)}"

        wb.save(path)
        messagebox.showinfo("Exportado", f"Archivo guardado en:\n{path}")

    # -----------------------------------------------------------------
    # Exportar a CSV
    # -----------------------------------------------------------------
    def export_csv(self, filtered_only=False):
        default_name = "Directorio_Negocios_filtrado.csv" if filtered_only else "Directorio_Negocios.csv"
        path = filedialog.asksaveasfilename(
            title="Guardar directorio de negocios (CSV)", defaultextension=".csv",
            initialfile=default_name, filetypes=[("CSV", "*.csv")]
        )
        if not path:
            return
        businesses, _grouped = self._grouped_for_export(filtered_only)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Rubro"] + self.EXPORT_HEADERS)
            for b in businesses:
                writer.writerow([b["category"]] + self._export_row(b))
        messagebox.showinfo("Exportado", f"Archivo guardado en:\n{path}")

    # -----------------------------------------------------------------
    # Exportar a PDF
    # -----------------------------------------------------------------
    def _build_pdf(self, path, filtered_only=False):
        _all_rows, grouped = self._grouped_for_export(filtered_only)
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                 leftMargin=1.2 * cm, rightMargin=1.2 * cm,
                                 topMargin=1.2 * cm, bottomMargin=1.2 * cm)
        title = "Directorio de Negocios" + (" (filtrado)" if filtered_only else "")
        elements = [Paragraph(title, styles["Title"]), Spacer(1, 10)]

        pdf_headers = ["Código", "Negocio", "Contacto", "Teléfono", "Correo", "Ciudad", "Estado", "Próx. seguim."]
        for name, meta in self.categories.items():
            rows = grouped.get(name, [])
            if filtered_only and not rows:
                continue
            elements.append(Paragraph(f"{name} ({meta['prefix']}) — {len(rows)} negocio(s)", styles["Heading2"]))
            table_data = [pdf_headers]
            for b in rows:
                table_data.append([b["code"], b["negocio"], b["contacto"], b["telefono"],
                                    b["correo"], b["ciudad"], b["estado"], b["proximo_contacto"]])
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E7EC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FA")]),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 16))

        doc.build(elements)

    def export_pdf(self, filtered_only=False):
        if not REPORTLAB_OK:
            messagebox.showerror(
                "Falta un paquete",
                "Para exportar a PDF necesitas instalar reportlab.\n\nAbre una terminal y ejecuta:\n\npip install reportlab"
            )
            return
        default_name = "Directorio_Negocios_filtrado.pdf" if filtered_only else "Directorio_Negocios.pdf"
        path = filedialog.asksaveasfilename(
            title="Guardar directorio de negocios (PDF)", defaultextension=".pdf",
            initialfile=default_name, filetypes=[("PDF", "*.pdf")]
        )
        if not path:
            return
        self._build_pdf(path, filtered_only=filtered_only)
        messagebox.showinfo("Exportado", f"Archivo guardado en:\n{path}")

    # -----------------------------------------------------------------
    # Imprimir listado directamente
    # -----------------------------------------------------------------
    def print_listing(self, filtered_only=False):
        if not REPORTLAB_OK:
            messagebox.showerror(
                "Falta un paquete",
                "Para imprimir necesitas instalar reportlab.\n\nAbre una terminal y ejecuta:\n\npip install reportlab"
            )
            return
        import tempfile
        tmp_path = os.path.join(
            tempfile.gettempdir(),
            f"listado_negocios_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        )
        try:
            self._build_pdf(tmp_path, filtered_only=filtered_only)
        except Exception as exc:
            messagebox.showerror("Error al preparar la impresión", f"No se pudo generar el PDF:\n{exc}")
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(tmp_path, "print")  # noqa: S606 — abre el diálogo de impresión de Windows
            else:
                webbrowser.open(tmp_path)
        except Exception:
            webbrowser.open(tmp_path)
        self.show_msg("Se generó el listado en PDF y se envió a tu impresora / visor predeterminado.")


if __name__ == "__main__":
    app = App()
    app.mainloop()
